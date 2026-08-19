"""Importación de contactos por Excel (.xlsx) y plantilla de guía.

Por qué existe: quien opera la plataforma no es técnica. Pedirle un CSV con la
columna literal `phone_e164` y el teléfono ya en formato E.164 es pedirle que
haga el trabajo de un programador. Aquí el contrato es al revés — el archivo se
descarga ya armado (`construir_plantilla`), los encabezados están en español y
el número se acepta como lo escribe la gente (`300 123-4567`, `57 300...`,
`+57 300 ...`) y se normaliza antes de tocar la base.

Tres reglas que este módulo respeta y que conviene no romper al editarlo:

- **Regla 1 (nada de PII en logs ni en mensajes de error).** El motivo de un
  rechazo dice `fila 7: el teléfono no tiene código de país`, nunca el número.
  El contenido del archivo no se loguea jamás, solo su tamaño.
- **El CHECK de la base manda.** `contacts.phone_e164` tiene
  `~ '^\\+[1-9][0-9]{6,18}$'` en Postgres. Todo lo que salga de
  `normalizar_telefono` cumple ese patrón o se rechaza aquí, con un motivo en
  español, antes de llegar al INSERT.
- **Idempotente por `(team_id, phone_e164)`.** Reimportar el mismo archivo no
  duplica: la segunda vez todo cae en "actualizados". Los atributos se
  **fusionan** (no se reemplazan): subir un Excel con solo la columna "Ciudad"
  no borra el atributo "Idioma" que el contacto ya tenía.
"""
from __future__ import annotations

import io
import logging
import re
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

from .. import models, schemas

logger = logging.getLogger(__name__)

# Tope de bytes del upload. El mismo que ya usaba el importador CSV.
MAX_EXCEL_BYTES = 2 * 1024 * 1024
# Tope de filas de datos. Más bajo que el del CSV a propósito: openpyxl
# materializa objetos por celda y un .xlsx de 2 MB comprimido puede traer
# cientos de miles de filas.
MAX_EXCEL_ROWS = 5000
# Tope de errores devueltos al cliente. Si el archivo entero está mal, la
# usuaria no necesita 5000 líneas: con las primeras entiende el problema.
MAX_ERRORES_REPORTADOS = 200

# Longitud máxima de la clave y del valor de un atributo. Evita que una celda
# con un párrafo entero termine en el JSONB.
MAX_LARGO_CLAVE_ATRIBUTO = 60
MAX_LARGO_VALOR_ATRIBUTO = 500
MAX_ATRIBUTOS_POR_CONTACTO = 40

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HOJA_DATOS = "Contactos"
HOJA_INSTRUCCIONES = "Instrucciones"

# Encabezados que se escriben en la plantilla. Los dos últimos son atributos
# de ejemplo: cualquier columna que no sea de las conocidas se guarda como
# atributo con el encabezado como nombre del campo.
ENCABEZADOS_PLANTILLA = [
    "Nombre",
    "Teléfono",
    "Correo",
    "Acepta mensajes",
    "Ciudad",
    "Idioma",
]

# Fila de ejemplo de la plantilla. Datos claramente ficticios (regla 8: este
# repo es público, aquí no va el teléfono ni el correo de nadie real).
FILA_EJEMPLO = ["María Gómez", "+57 300 000 0000", "maria@ejemplo.com", "sí", "Cali", "Español"]
TELEFONO_EJEMPLO_E164 = "+573000000000"


# ---------------------------------------------------------------------------
# Encabezados
# ---------------------------------------------------------------------------

def _sin_acentos(texto: str) -> str:
    descompuesto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in descompuesto if unicodedata.category(c) != "Mn")


def normalizar_encabezado(bruto: Any) -> str:
    """`  Teléfono  ` → `telefono`. Sirve para comparar contra los alias."""
    if bruto is None:
        return ""
    texto = _sin_acentos(str(bruto)).lower().replace("_", " ").replace("-", " ")
    return re.sub(r"\s+", " ", texto).strip()


_ALIAS_TELEFONO = {
    "telefono", "telefono whatsapp", "telefono celular", "celular", "movil",
    "whatsapp", "numero", "numero de whatsapp", "numero de telefono",
    "phone", "phone e164", "phone number", "tel",
}
_ALIAS_NOMBRE = {"nombre", "nombre completo", "nombres", "name", "full name", "contacto"}
_ALIAS_EMAIL = {"correo", "correo electronico", "email", "e mail", "mail"}
_ALIAS_OPT_IN = {
    "acepta mensajes", "acepta", "acepta whatsapp", "autoriza", "autorizacion",
    "consentimiento", "opt in", "optin", "suscrito",
}
_ALIAS_OPT_IN_SOURCE = {
    "origen del permiso", "origen", "fuente", "fuente del permiso",
    "opt in source", "optin source",
}

# Claves canónicas de las columnas conocidas.
COL_TELEFONO = "phone"
COL_NOMBRE = "name"
COL_EMAIL = "email"
COL_OPT_IN = "opt_in"
COL_OPT_IN_SOURCE = "opt_in_source"


def clasificar_encabezado(bruto: Any) -> Optional[str]:
    """Devuelve la clave canónica de la columna, o None si es un atributo."""
    norm = normalizar_encabezado(bruto)
    if not norm:
        return None
    if norm in _ALIAS_TELEFONO:
        return COL_TELEFONO
    if norm in _ALIAS_NOMBRE:
        return COL_NOMBRE
    if norm in _ALIAS_EMAIL:
        return COL_EMAIL
    if norm in _ALIAS_OPT_IN:
        return COL_OPT_IN
    if norm in _ALIAS_OPT_IN_SOURCE:
        return COL_OPT_IN_SOURCE
    return None


# ---------------------------------------------------------------------------
# Celdas
# ---------------------------------------------------------------------------

def _texto_de_celda(valor: Any) -> str:
    """Convierte lo que devuelve openpyxl a texto plano y usable.

    Excel guarda los números como float: una columna de teléfonos escrita sin
    `+` llega como `5.73001234567e+11`. Sin este paso el importador rechazaría
    archivos perfectamente válidos.
    """
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "sí" if valor else "no"
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return repr(valor)
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    return str(valor).strip()


_VERDADEROS = {"si", "s", "sí", "yes", "y", "true", "verdadero", "1", "x", "ok", "acepta"}
_FALSOS = {"no", "n", "false", "falso", "0", "-", "rechaza"}


def parsear_booleano(bruto: Any) -> Optional[bool]:
    """`sí`/`no`/`1`/`0`/`true`… → bool. Vacío o ilegible → None (sin opinión)."""
    texto = normalizar_encabezado(_texto_de_celda(bruto))
    if not texto:
        return None
    if texto in _VERDADEROS:
        return True
    if texto in _FALSOS:
        return False
    return None


# ---------------------------------------------------------------------------
# Teléfonos
# ---------------------------------------------------------------------------

# El mismo patrón que el CHECK `ck_contacts_phone_e164` de la tabla.
E164_RE = re.compile(r"^\+[1-9][0-9]{6,18}$")

# Caracteres de "adorno" que la gente escribe dentro de un teléfono.
_BASURA_TELEFONO = re.compile(r"[\s\-\.\(\)/–—_ ​]")


def normalizar_telefono(
    bruto: Any, prefijo_pais: Optional[str] = None
) -> Tuple[Optional[str], Optional[str]]:
    """`(300) 123-4567` + país `57` → `+573001234567`.

    Devuelve `(e164, None)` si quedó válido, o `(None, motivo)` con un motivo
    en español apto para mostrarle a la usuaria. El motivo NUNCA incluye el
    número (regla 1): el propósito es que ella encuentre la fila, no que el
    dato viaje por logs o pantallas.
    """
    texto = _texto_de_celda(bruto).strip()
    if not texto:
        return None, "falta el teléfono"

    limpio = _BASURA_TELEFONO.sub("", texto)
    # Un `+` solo se acepta al inicio; en medio es un error de tipeo.
    if "+" in limpio[1:]:
        return None, "el teléfono tiene caracteres que no son números"

    tiene_mas = limpio.startswith("+")
    cuerpo = limpio[1:] if tiene_mas else limpio

    # `00` al inicio es el prefijo internacional de marcación (00 57 300…).
    if not tiene_mas and cuerpo.startswith("00"):
        cuerpo = cuerpo[2:]
        tiene_mas = True

    if not cuerpo:
        return None, "falta el teléfono"
    if not cuerpo.isdigit():
        return None, "el teléfono tiene caracteres que no son números"

    if not tiene_mas:
        prefijo = (prefijo_pais or "").strip().lstrip("+")
        if len(cuerpo) <= 10:
            # Un número de 10 dígitos o menos no trae código de país. Si no
            # hay país por defecto configurado, prefijar un `+` a secas
            # produciría un contacto inalcanzable — y peor: sin error visible.
            if not prefijo:
                return None, (
                    "el teléfono no tiene código de país "
                    "(escríbelo como +57… o elige un país por defecto al importar)"
                )
            cuerpo = prefijo + cuerpo
        # Con más de 10 dígitos se asume que el código de país ya viene dentro.

    candidato = "+" + cuerpo
    if not E164_RE.match(candidato):
        if len(cuerpo) < 7:
            return None, "el teléfono tiene muy pocos dígitos"
        if len(cuerpo) > 19:
            return None, "el teléfono tiene demasiados dígitos"
        if cuerpo.startswith("0"):
            return None, "el código de país no puede empezar en 0"
        return None, "el teléfono no tiene un formato válido"
    return candidato, None


# ---------------------------------------------------------------------------
# Plantilla de guía
# ---------------------------------------------------------------------------

INSTRUCCIONES: List[Tuple[str, str]] = [
    (
        "1. Llena la hoja «Contactos»",
        "Una fila por persona. La fila 2 es un ejemplo: bórrala o déjala, el "
        "sistema la ignora automáticamente.",
    ),
    (
        "2. El teléfono es lo único obligatorio",
        "Escríbelo con el código de país: +57 300 000 0000. También se acepta "
        "sin el «+», con espacios o con guiones. Si tus números no llevan "
        "código de país, elige el país por defecto en la pantalla de "
        "importación y el sistema lo agrega.",
    ),
    (
        "3. «Acepta mensajes»",
        "Escribe «sí» si la persona autorizó recibir mensajes de WhatsApp, o "
        "«no» si no lo hizo. Si lo dejas vacío, se guarda como «sí» para los "
        "contactos nuevos y no se cambia el valor de los que ya existían. "
        "A quien tenga «no» nunca se le envía una campaña.",
    ),
    (
        "4. Agrega las columnas que quieras",
        "Cualquier columna extra (Ciudad, Idioma, Destino favorito, Cumpleaños…) "
        "se guarda como un dato del contacto y después la puedes usar para "
        "personalizar el mensaje de una campaña. El título de la columna es el "
        "nombre del dato: escríbelo como quieres verlo.",
    ),
    (
        "5. Volver a subir el mismo archivo no duplica",
        "Los contactos se identifican por el teléfono. Si ya existe, se "
        "actualiza; si no, se crea. Puedes corregir el archivo y volverlo a "
        "subir las veces que necesites.",
    ),
    (
        "6. No cambies el nombre de la hoja «Contactos»",
        "Si la renombras, el sistema usa la primera hoja del archivo.",
    ),
]


def construir_plantilla() -> bytes:
    """Arma el .xlsx de guía: hoja de datos + hoja de instrucciones."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    hoja = wb.active
    hoja.title = HOJA_DATOS

    # Paleta Gloma: Deep Forest de fondo, texto claro.
    relleno = PatternFill("solid", fgColor="FF004D40")
    fuente_encabezado = Font(bold=True, color="FFF5FAF9", size=11)

    hoja.append(ENCABEZADOS_PLANTILLA)
    for idx, _ in enumerate(ENCABEZADOS_PLANTILLA, start=1):
        celda = hoja.cell(row=1, column=idx)
        celda.fill = relleno
        celda.font = fuente_encabezado
        celda.alignment = Alignment(horizontal="left", vertical="center")
        hoja.column_dimensions[get_column_letter(idx)].width = 24
    hoja.row_dimensions[1].height = 22

    hoja.append(FILA_EJEMPLO)
    for idx, _ in enumerate(FILA_EJEMPLO, start=1):
        hoja.cell(row=2, column=idx).font = Font(italic=True, color="FF4A7A72")

    hoja.cell(row=4, column=1).value = (
        "↑ La fila 2 es un ejemplo y el sistema la ignora. "
        "Escribe tus contactos desde la fila 3 (o reemplaza el ejemplo)."
    )
    hoja.cell(row=4, column=1).font = Font(italic=True, color="FF4A7A72", size=10)
    # El teléfono se escribe como texto: si Excel lo trata como número, se
    # come el «+» y los ceros a la izquierda.
    for fila in range(2, 500):
        hoja.cell(row=fila, column=2).number_format = "@"
    hoja.freeze_panes = "A2"

    guia = wb.create_sheet(HOJA_INSTRUCCIONES)
    guia.column_dimensions["A"].width = 44
    guia.column_dimensions["B"].width = 96
    guia.cell(row=1, column=1).value = "Cómo llenar esta plantilla"
    guia.cell(row=1, column=1).font = Font(bold=True, size=14, color="FF004D40")
    fila = 3
    for titulo, detalle in INSTRUCCIONES:
        guia.cell(row=fila, column=1).value = titulo
        guia.cell(row=fila, column=1).font = Font(bold=True, color="FF003A30")
        guia.cell(row=fila, column=1).alignment = Alignment(vertical="top")
        guia.cell(row=fila, column=2).value = detalle
        guia.cell(row=fila, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        guia.row_dimensions[fila].height = 34
        fila += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Importación
# ---------------------------------------------------------------------------

def _limpiar_atributo(clave: Any, valor: Any) -> Optional[Tuple[str, str]]:
    nombre = str(clave).strip()[:MAX_LARGO_CLAVE_ATRIBUTO] if clave is not None else ""
    if not nombre:
        return None
    texto = _texto_de_celda(valor).strip()
    if not texto:
        return None
    return nombre, texto[:MAX_LARGO_VALOR_ATRIBUTO]


def _upsert_contacto(
    db: Session,
    team_id: int,
    *,
    phone_e164: str,
    name: Optional[str],
    email: Optional[str],
    opt_in: Optional[bool],
    atributos: Dict[str, str],
) -> bool:
    """Crea o actualiza por `(team_id, phone_e164)`. Devuelve True si fue alta.

    Los atributos se fusionan con los que el contacto ya tenía: una columna
    ausente en el Excel no borra un dato guardado antes. Para *quitar* un
    atributo está la pantalla de edición del contacto (PATCH), que sí manda el
    diccionario completo.
    """
    existente = (
        db.query(models.Contact)
        .filter(
            models.Contact.team_id == team_id,
            models.Contact.phone_e164 == phone_e164,
        )
        .first()
    )

    if existente is not None:
        if name:
            existente.name = name
        if email:
            existente.email = email
        if opt_in is not None:
            existente.opt_in = bool(opt_in)
        if atributos:
            fusionados = dict(existente.attributes or {})
            fusionados.update(atributos)
            # Reasignar el dict (no mutarlo) es lo que hace que SQLAlchemy
            # marque la columna JSONB como sucia.
            existente.attributes = dict(
                list(fusionados.items())[:MAX_ATRIBUTOS_POR_CONTACTO]
            )
        existente.updated_at = datetime.utcnow()
        db.commit()
        return False

    contacto = models.Contact(
        team_id=team_id,
        phone_e164=phone_e164,
        name=name,
        email=email,
        attributes=dict(list(atributos.items())[:MAX_ATRIBUTOS_POR_CONTACTO]),
        opt_in=True if opt_in is None else bool(opt_in),
        opt_in_source="import_excel",
    )
    db.add(contacto)
    db.commit()
    return True


def _abrir_libro(contenido: bytes):
    from openpyxl import load_workbook

    libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    if HOJA_DATOS in libro.sheetnames:
        return libro, libro[HOJA_DATOS]
    return libro, libro[libro.sheetnames[0]]


def importar(
    db: Session,
    team_id: int,
    contenido: bytes,
    prefijo_pais: Optional[str] = None,
) -> schemas.ContactExcelImportResult:
    """Lee el .xlsx y hace el alta/actualización masiva.

    No lanza por una fila mala: cada rechazo se acumula con su número de fila
    y su motivo, y la importación sigue. Solo un archivo ilegible (no es un
    .xlsx, viene corrupto, no tiene columna de teléfono) corta el proceso, y
    aun así devuelve un resultado con el motivo, no una excepción.
    """
    errores: List[schemas.ContactExcelRowError] = []

    def _vacio(motivo: str) -> schemas.ContactExcelImportResult:
        return schemas.ContactExcelImportResult(
            total=0, created=0, updated=0, rejected=0,
            errors=[], detected_attributes=[], notice=motivo,
        )

    try:
        libro, hoja = _abrir_libro(contenido)
    except Exception:
        # El detalle (traceback de openpyxl) se queda del lado del servidor.
        logger.exception("import_excel: archivo ilegible (team_id=%s)", team_id)
        return _vacio(
            "No se pudo leer el archivo. Asegúrate de subir un .xlsx "
            "(Excel) generado desde la plantilla."
        )

    try:
        filas = hoja.iter_rows(values_only=True)
        try:
            encabezados = next(filas)
        except StopIteration:
            return _vacio("El archivo está vacío: no tiene ni encabezados.")

        # Mapa de columna → destino. Las que no son conocidas van a atributos.
        canonicas: Dict[int, str] = {}
        atributos_col: Dict[int, str] = {}
        for idx, bruto in enumerate(encabezados or ()):
            clave = clasificar_encabezado(bruto)
            if clave is not None:
                # Si el usuario duplicó una columna conocida, gana la primera.
                if clave not in canonicas.values():
                    canonicas[idx] = clave
                continue
            nombre = str(bruto).strip()[:MAX_LARGO_CLAVE_ATRIBUTO] if bruto else ""
            if nombre:
                atributos_col[idx] = nombre

        if COL_TELEFONO not in canonicas.values():
            return _vacio(
                "El archivo no tiene una columna de teléfono. Descarga la "
                "plantilla y usa el encabezado «Teléfono»."
            )

        col_telefono = next(i for i, k in canonicas.items() if k == COL_TELEFONO)
        col_nombre = next((i for i, k in canonicas.items() if k == COL_NOMBRE), None)
        col_email = next((i for i, k in canonicas.items() if k == COL_EMAIL), None)
        col_opt_in = next((i for i, k in canonicas.items() if k == COL_OPT_IN), None)

        total = 0
        creados = 0
        actualizados = 0
        rechazados = 0
        atributos_vistos: List[str] = []

        for offset, fila in enumerate(filas):
            numero_fila = offset + 2  # la 1 es el encabezado
            if fila is None or all(_texto_de_celda(v) == "" for v in fila):
                continue  # fila en blanco: ni se cuenta ni molesta

            total += 1
            if total > MAX_EXCEL_ROWS:
                errores.append(
                    schemas.ContactExcelRowError(
                        row=numero_fila,
                        reason=(
                            f"el archivo supera el máximo de {MAX_EXCEL_ROWS} "
                            "filas; el resto no se procesó"
                        ),
                    )
                )
                total -= 1
                break

            def _celda(idx: Optional[int]) -> Any:
                if idx is None or idx >= len(fila):
                    return None
                return fila[idx]

            telefono, motivo = normalizar_telefono(
                _celda(col_telefono), prefijo_pais
            )
            if telefono is None:
                rechazados += 1
                if len(errores) < MAX_ERRORES_REPORTADOS:
                    errores.append(
                        schemas.ContactExcelRowError(
                            row=numero_fila, reason=motivo or "teléfono inválido"
                        )
                    )
                continue

            if telefono == TELEFONO_EJEMPLO_E164:
                rechazados += 1
                if len(errores) < MAX_ERRORES_REPORTADOS:
                    errores.append(
                        schemas.ContactExcelRowError(
                            row=numero_fila,
                            reason=(
                                "es la fila de ejemplo de la plantilla, no se "
                                "importó (puedes borrarla del archivo)"
                            ),
                        )
                    )
                continue

            nombre = _texto_de_celda(_celda(col_nombre)).strip()[:120] or None
            correo_bruto = _texto_de_celda(_celda(col_email)).strip()
            correo = correo_bruto[:255] or None
            if correo and ("@" not in correo or " " in correo):
                rechazados += 1
                if len(errores) < MAX_ERRORES_REPORTADOS:
                    errores.append(
                        schemas.ContactExcelRowError(
                            row=numero_fila,
                            reason="el correo no tiene un formato válido",
                        )
                    )
                continue
            opt_in = parsear_booleano(_celda(col_opt_in)) if col_opt_in is not None else None

            atributos: Dict[str, str] = {}
            for idx, nombre_attr in atributos_col.items():
                par = _limpiar_atributo(nombre_attr, _celda(idx))
                if par is None:
                    continue
                atributos[par[0]] = par[1]
                if par[0] not in atributos_vistos:
                    atributos_vistos.append(par[0])

            try:
                fue_alta = _upsert_contacto(
                    db,
                    team_id,
                    phone_e164=telefono,
                    name=nombre,
                    email=correo,
                    opt_in=opt_in,
                    atributos=atributos,
                )
            except Exception:
                db.rollback()
                # Ni el teléfono ni la fila cruda entran al log (regla 1).
                logger.exception(
                    "import_excel: error persistiendo fila (team_id=%s, fila=%s)",
                    team_id,
                    numero_fila,
                )
                rechazados += 1
                if len(errores) < MAX_ERRORES_REPORTADOS:
                    errores.append(
                        schemas.ContactExcelRowError(
                            row=numero_fila,
                            reason="no se pudo guardar este contacto, inténtalo de nuevo",
                        )
                    )
                continue

            if fue_alta:
                creados += 1
            else:
                actualizados += 1
    finally:
        try:
            libro.close()
        except Exception:  # pragma: no cover - cierre best-effort
            pass

    aviso = None
    if rechazados and not creados and not actualizados:
        aviso = (
            "Ninguna fila se pudo importar. Revisa la columna de teléfono: "
            "debe incluir el código de país (+57, +52…)."
        )
    elif len(errores) >= MAX_ERRORES_REPORTADOS:
        aviso = (
            f"Se muestran los primeros {MAX_ERRORES_REPORTADOS} errores. "
            "Corrige esos y vuelve a subir el archivo."
        )

    return schemas.ContactExcelImportResult(
        total=total,
        created=creados,
        updated=actualizados,
        rejected=rechazados,
        errors=errores,
        detected_attributes=atributos_vistos,
        notice=aviso,
    )


# ---------------------------------------------------------------------------
# Catálogo de campos (para el selector de variables de campaña)
# ---------------------------------------------------------------------------

# Tope de contactos que se escanean para descubrir qué atributos existen. Con
# 2000 filas ya se vio cualquier columna que la usuaria haya subido; ir más
# allá solo alarga la query.
MAX_CONTACTOS_ESCANEADOS = 2000


def campos_disponibles(db: Session, team_id: int) -> schemas.ContactFieldsOut:
    """Qué campos se pueden usar para personalizar un mensaje.

    Devuelve el catálogo, NO los valores: aquí no viaja PII (regla 2). Los
    `token` son la convención que se guarda en `campaigns.template_variables_json`
    cuando la usuaria elige "usar el dato del contacto" en el asistente.
    """
    campos = [
        schemas.ContactFieldOut(
            key="name",
            label="Nombre del contacto",
            token="{{contact.name}}",
            source="base",
            contacts=0,
        ),
        schemas.ContactFieldOut(
            key="phone_e164",
            label="Teléfono del contacto",
            token="{{contact.phone}}",
            source="base",
            contacts=0,
        ),
    ]

    filas = (
        db.query(models.Contact.attributes)
        .filter(models.Contact.team_id == team_id)
        .order_by(models.Contact.updated_at.desc())
        .limit(MAX_CONTACTOS_ESCANEADOS)
        .all()
    )

    conteo: Dict[str, int] = {}
    for (attrs,) in filas:
        if not isinstance(attrs, dict):
            continue
        for clave in attrs:
            nombre = str(clave).strip()
            if nombre:
                conteo[nombre] = conteo.get(nombre, 0) + 1

    for nombre in sorted(conteo, key=lambda k: (-conteo[k], k.lower())):
        campos.append(
            schemas.ContactFieldOut(
                key=nombre,
                label=nombre,
                token="{{contact.attributes." + nombre + "}}",
                source="attribute",
                contacts=conteo[nombre],
            )
        )

    return schemas.ContactFieldsOut(fields=campos, scanned_contacts=len(filas))
