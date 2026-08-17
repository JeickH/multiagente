"""Sprint "Ayuda a Cali": reportes de mascotas perdidas y encontradas.

Concentra todo lo que el bot y el panel necesitan sobre `mascotas`:

  - **Storage de fotos** por reporte, en `mascotas/<codigo>/`. Dos backends
    intercambiables: S3 (producción, `MASCOTAS_BUCKET`) y filesystem local
    (desarrollo). Las fotos NUNCA se sirven desde el bucket: el bucket es
    privado y el backend hace de proxy (`GET /mascotas/foto/...`), así no hay
    que abrir acceso público a S3.
  - **Adopción de fotos**: el ciudadano manda las fotos antes de que el bot
    termine de recoger los datos, así que se suben contra un `upload_session`
    (uuid del chat) y se mueven a la carpeta del reporte cuando este se crea.
  - **Búsqueda por coincidencias**: scoring campo a campo (nadie recuerda todos
    los datos), sin exigir que coincida todo. Cruza 'perdida' ↔ 'encontrada'.
  - **Exportación a Excel** del listado vigente.

Regla de privacidad del módulo: el teléfono de contacto SOLO se entrega cuando
la persona confirma que reconoce a la mascota. Ni la búsqueda ni la ficha lo
incluyen.
"""
from __future__ import annotations

import io
import json
import logging
import os
import re
import unicodedata
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import func
from sqlalchemy.orm import Session

from .. import models
from . import imagenes

logger = logging.getLogger(__name__)

# Formatos aceptados en el chat (lo que un celular manda al compartir una foto).
ALLOWED_IMAGE_TYPES = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/heic": ".heic",
}
MAX_PHOTO_BYTES = 8 * 1024 * 1024   # 8 MB por foto
MAX_FOTOS_POR_REPORTE = 6

_LOCAL_MEDIA_ROOT = Path(os.getenv("MASCOTAS_MEDIA_DIR", "/app/media"))

# Plataformas hermanas de las que importamos reportes. La clave es el `source`
# de la fila; el valor, cómo se le nombra a la persona en el chat y el panel.
ORIGEN_NOMBRES = {
    "mascotasporcolombia": "Mascotas por Colombia",
    "patitasacasa": "Patitas a Casa",
    "petsearch": "PetSearch Colombia",
    "encontradogs": "encontradogs",
    "proteccionanimal": "Protección y Bienestar Animal del Valle del Cauca",
    "royipets": "RoyiPets",
}


# ---------------------------------------------------------------------------
# Storage
# ---------------------------------------------------------------------------

def _bucket() -> str:
    return (os.getenv("MASCOTAS_BUCKET") or "").strip()


def _s3():
    import boto3  # import perezoso: en local el backend corre sin S3

    return boto3.client("s3", region_name=os.getenv("AWS_REGION", "sa-east-1"))


def _put_object(
    key: str, data: bytes, content_type: str, metadata: Optional[Dict[str, str]] = None
) -> None:
    bucket = _bucket()
    if bucket:
        _s3().put_object(
            Bucket=bucket, Key=key, Body=data, ContentType=content_type,
            Metadata=metadata or {},
        )
        return
    path = _LOCAL_MEDIA_ROOT / key
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def storage_uri(key: str) -> str:
    """Dónde vive físicamente una foto, para mostrarlo en el panel.

    `s3://bucket/mascotas/<codigo>/<archivo>` en producción; `file://…` cuando
    el backend corre sin bucket (desarrollo).
    """
    bucket = _bucket()
    if bucket:
        return f"s3://{bucket}/{key}"
    return f"file://{_LOCAL_MEDIA_ROOT / key}"


def _get_object(key: str) -> Optional[bytes]:
    bucket = _bucket()
    if bucket:
        try:
            resp = _s3().get_object(Bucket=bucket, Key=key)
            return resp["Body"].read()
        except Exception:
            logger.exception("mascotas: no se pudo leer %s de S3", key)
            return None
    path = _LOCAL_MEDIA_ROOT / key
    try:
        return path.read_bytes()
    except OSError:
        return None


def _delete_object(key: str) -> bool:
    """Borra un archivo del storage. False si no se pudo (la fila se borra
    igual: un objeto huérfano molesta menos que un registro fantasma)."""
    bucket = _bucket()
    try:
        if bucket:
            _s3().delete_object(Bucket=bucket, Key=key)
            return True
        (_LOCAL_MEDIA_ROOT / key).unlink(missing_ok=True)
        return True
    except Exception:
        logger.exception("mascotas: no se pudo borrar %s del storage", key)
        return False


def _move_object(src_key: str, dst_key: str) -> bool:
    """Mueve un objeto dentro del storage. False si no se pudo (el llamador
    conserva la clave vieja: perder la foto es peor que tenerla mal ubicada)."""
    if src_key == dst_key:
        return True
    bucket = _bucket()
    try:
        if bucket:
            _s3().copy_object(
                Bucket=bucket, Key=dst_key, CopySource={"Bucket": bucket, "Key": src_key}
            )
            _s3().delete_object(Bucket=bucket, Key=src_key)
            return True
        src, dst = _LOCAL_MEDIA_ROOT / src_key, _LOCAL_MEDIA_ROOT / dst_key
        dst.parent.mkdir(parents=True, exist_ok=True)
        src.replace(dst)
        return True
    except Exception:
        logger.exception("mascotas: no se pudo mover %s -> %s", src_key, dst_key)
        return False


def guardar_foto(
    db: Session,
    data: bytes,
    content_type: str,
    upload_session: Optional[str] = None,
    mascota: Optional[models.Mascota] = None,
) -> models.MascotaFoto:
    """Sube una foto al storage y la registra.

    Sin `mascota` la foto queda en el limbo (`pendientes/<upload_session>/`)
    hasta que `adoptar_fotos()` la mueva al reporte.

    La foto se comprime antes de guardarla (#360): llegaban de 4 MB desde el
    celular y el sitio las servía tal cual. Si la compresión falla o no gana
    nada, se guarda el original y la foto queda sin marcar, para que el barrido
    del bucket la tome después.
    """
    bytes_original = len(data)
    comprimida = imagenes.comprimir(data)
    if comprimida is not None:
        data, content_type = comprimida, "image/jpeg"

    ext = ALLOWED_IMAGE_TYPES.get(content_type, ".jpg")
    nombre = f"{uuid.uuid4().hex}{ext}"
    carpeta = f"mascotas/{mascota.codigo}" if mascota else f"pendientes/{upload_session}"
    key = f"{carpeta}/{nombre}"
    # La marca en el objeto es lo que hace que el barrido del bucket ni siquiera
    # se moleste en bajar esta foto: ya viene comprimida de aquí.
    _put_object(
        key, data, content_type,
        metadata={"optimizado": imagenes.MARCA, "bytes-original": str(bytes_original)}
        if comprimida is not None else None,
    )
    if comprimida is not None:
        logger.info(
            "mascotas: foto comprimida %d -> %d bytes (%.0f%% menos)",
            bytes_original, len(data), 100 * (1 - len(data) / bytes_original),
        )

    foto = models.MascotaFoto(
        mascota_id=mascota.id if mascota else None,
        upload_session=None if mascota else upload_session,
        storage_key=key,
        content_type=content_type,
        bytes_size=len(data),
        optimizada=comprimida is not None,
        optimizada_at=datetime.utcnow() if comprimida is not None else None,
        bytes_original=bytes_original,
    )
    db.add(foto)
    db.commit()
    db.refresh(foto)
    return foto


def adoptar_fotos(db: Session, mascota: models.Mascota, upload_session: str) -> int:
    """Asocia al reporte las fotos que el ciudadano subió durante el chat."""
    if not upload_session:
        return 0
    pendientes = (
        db.query(models.MascotaFoto)
        .filter(
            models.MascotaFoto.upload_session == upload_session,
            models.MascotaFoto.mascota_id.is_(None),
        )
        .order_by(models.MascotaFoto.id)
        .limit(MAX_FOTOS_POR_REPORTE)
        .all()
    )
    for foto in pendientes:
        nuevo_key = f"mascotas/{mascota.codigo}/{foto.storage_key.rsplit('/', 1)[-1]}"
        if _move_object(foto.storage_key, nuevo_key):
            foto.storage_key = nuevo_key
        foto.mascota_id = mascota.id
        foto.upload_session = None
    if pendientes:
        db.commit()
    return len(pendientes)


def contar_fotos_pendientes(db: Session, upload_session: str) -> int:
    if not upload_session:
        return 0
    return (
        db.query(models.MascotaFoto)
        .filter(
            models.MascotaFoto.upload_session == upload_session,
            models.MascotaFoto.mascota_id.is_(None),
        )
        .count()
    )


def leer_foto(db: Session, mascota_codigo: str, foto_id: int) -> Optional[Tuple[bytes, str]]:
    """Bytes + content-type de una foto, validando que pertenezca al reporte."""
    foto = (
        db.query(models.MascotaFoto)
        .join(models.Mascota, models.MascotaFoto.mascota_id == models.Mascota.id)
        .filter(
            models.MascotaFoto.id == foto_id,
            models.Mascota.codigo == mascota_codigo,
        )
        .first()
    )
    if foto is None:
        return None
    data = _get_object(foto.storage_key)
    if data is None:
        return None
    return data, foto.content_type or "image/jpeg"


def leer_foto_pendiente(
    db: Session, upload_session: str
) -> Optional[Tuple[bytes, str]]:
    """La primera foto que la persona adjuntó en el chat, antes de que exista
    un reporte al cual colgarla.

    Es la que sirve para comparar visualmente contra las candidatas: en ese
    momento la persona ya mostró a su mascota pero todavía no registró nada.
    """
    if not upload_session:
        return None
    foto = (
        db.query(models.MascotaFoto)
        .filter(
            models.MascotaFoto.upload_session == upload_session,
            models.MascotaFoto.mascota_id.is_(None),
        )
        .order_by(models.MascotaFoto.id)
        .first()
    )
    if foto is None:
        return None
    data = _get_object(foto.storage_key)
    if data is None:
        return None
    return data, foto.content_type or "image/jpeg"


def leer_primera_foto(
    db: Session, codigo: str
) -> Optional[Tuple[bytes, str]]:
    """La primera foto de un reporte, para comparar contra otra."""
    mascota = obtener(db, codigo)
    fotos = list(getattr(mascota, "fotos", None) or [])
    if not fotos:
        return None
    data = _get_object(fotos[0].storage_key)
    if data is None:
        return None
    return data, fotos[0].content_type or "image/jpeg"


# ---------------------------------------------------------------------------
# Normalización y scoring
# ---------------------------------------------------------------------------

_PALABRAS_VACIAS = {
    "de", "la", "el", "los", "las", "un", "una", "y", "o", "en", "con", "por",
    "para", "del", "al", "es", "se", "que", "mi", "su", "muy", "cerca", "barrio",
    "calle", "carrera", "avenida", "cra", "cll", "av", "no", "sin", "color",
    "perro", "perra", "gato", "gata", "mascota", "raza", "aproximadamente",
}


def normalizar(texto: Optional[str]) -> str:
    """minúsculas, sin tildes y sin puntuación — para comparar lo que escribe
    la gente ('Café', 'cafe', 'CAFÉ' son el mismo color)."""
    if not texto:
        return ""
    limpio = unicodedata.normalize("NFKD", str(texto))
    limpio = "".join(c for c in limpio if not unicodedata.combining(c))
    limpio = re.sub(r"[^a-zA-Z0-9ñÑ\s]", " ", limpio.lower())
    return re.sub(r"\s+", " ", limpio).strip()


def _tokens(texto: Optional[str]) -> set:
    """Palabras con contenido, ya canonizadas (criollo/mestizo → criollo)."""
    return {
        _CANONICO.get(t, t)
        for t in normalizar(texto).split()
        if len(t) > 2 and t not in _PALABRAS_VACIAS
    }


# Palabras distintas que la gente usa para lo mismo. Cada grupo se colapsa a su
# primer término antes de comparar, así "criollo" y "mestizo" hacen match — que
# es como se describe a la mayoría de los animales de la calle.
_SINONIMOS = (
    ("criollo", "mestizo", "mestiza", "criolla", "callejero", "callejera",
     "sin raza", "corriente", "comun", "chandoso", "chanda"),
    ("cafe", "marron", "chocolate", "carmelito", "carmelita"),
    ("dorado", "amarillo", "beige", "crema", "mostaza", "rubio"),
    ("gris", "plateado", "azul"),
    ("atigrado", "rayado", "tigrillo", "tigre"),
    ("manchado", "pintado", "moteado"),
    ("pequeno", "chico", "chiquito", "small", "mini"),
    ("grande", "big", "gigante"),
    ("cachorro", "cachorra", "bebe", "puppy", "joven"),
)

# Índice palabra → término canónico del grupo.
_CANONICO = {
    palabra: grupo[0] for grupo in _SINONIMOS for palabra in grupo
}


def _canonizar(texto: str) -> str:
    """Reemplaza sinónimos por su término canónico. Recibe texto ya normalizado."""
    if not texto:
        return texto
    # Primero las expresiones de varias palabras ("sin raza"), luego palabra a
    # palabra: si no se hace en ese orden, "sin raza" nunca coincide.
    for palabra, canonico in _CANONICO.items():
        if " " in palabra and palabra in texto:
            texto = texto.replace(palabra, canonico)
    return " ".join(_CANONICO.get(p, p) for p in texto.split())


# Ciudades y regiones que comparte casi toda la base: que dos reportes digan
# "Cali" no acerca en nada, porque lo dicen todos. Solo el barrio discrimina.
_ZONAS_GENERICAS = frozenset({
    "cali", "valle", "cauca", "colombia", "palmira", "medellin", "bogota",
    "pereira", "armenia", "manizales", "sin", "ubicacion", "precisa",
})


def _score_texto(
    consulta: Optional[str],
    valor: Optional[str],
    peso: int,
    vacios: frozenset = frozenset(),
) -> int:
    """Puntaje de un campo de texto libre: exacto vale el peso completo,
    contenido o tokens compartidos valen menos. Sin dato no resta.

    `vacios` son términos que no discriminan (la ciudad, en la zona): si lo
    único que comparten los dos textos está en esa lista, el puntaje es 0.
    """
    a, b = _canonizar(normalizar(consulta)), _canonizar(normalizar(valor))
    if not a or not b:
        return 0

    if vacios:
        # Si al quitar los términos genéricos no queda nada en común, no suma.
        utiles_a, utiles_b = _tokens(a) - vacios, _tokens(b) - vacios
        if not (utiles_a & utiles_b):
            return 0

    if a == b:
        return peso
    if a in b or b in a:
        return max(1, peso - 1)
    comunes = (_tokens(consulta) & _tokens(valor)) - vacios
    if comunes:
        return max(1, peso - 2)
    return 0


ESPECIES = ("perro", "gato", "otra")


def normalizar_especie(valor: Optional[str]) -> str:
    """'perrito', 'canino', 'cachorro' → 'perro'. Desconocido → 'otra'."""
    v = normalizar(valor)
    if not v:
        return ""
    if any(k in v for k in ("perr", "canin", "cachorr", "peluch")):
        return "perro"
    if any(k in v for k in ("gat", "felin", "michi", "minin")):
        return "gato"
    return "otra"


def normalizar_sexo(valor: Optional[str]) -> Optional[str]:
    v = normalizar(valor)
    if not v:
        return None
    if v.startswith("m") or "macho" in v:
        return "macho"
    if v.startswith("h") or "hembra" in v:
        return "hembra"
    return "desconocido"


def _evaluar(criterios: Dict[str, Any], m: models.Mascota) -> Tuple[int, Dict[str, int]]:
    """Cuánto se parece un reporte a lo que describe la persona, y por qué.

    Filosofía: **nada es obligatorio**. Cada dato que coincide suma; los datos
    que la persona no sabe simplemente no puntúan. La especie sí es un filtro
    duro: un perro nunca es un gato.

    Devuelve (puntaje, desglose por campo). El desglose es lo que el panel le
    muestra al equipo para que entienda una coincidencia antes de llamar.
    """
    especie = normalizar_especie(criterios.get("especie"))
    if especie and m.especie and especie != m.especie:
        return -1, {}   # descartado

    detalle: Dict[str, int] = {}

    def _sumar(campo: str, puntos: int) -> None:
        if puntos > 0:
            detalle[campo] = puntos

    if especie and especie == m.especie:
        _sumar("especie", 2)
    # El peso está en lo FÍSICO y en la zona, no en el nombre: quien encuentra
    # una mascota casi nunca sabe cómo se llama, y dos perros del barrio pueden
    # llamarse igual. El nombre vale como desempate, nada más.
    _sumar("raza", _score_texto(criterios.get("raza"), m.raza, 5))
    _sumar("color", _score_texto(criterios.get("color"), m.color, 5))
    _sumar("tamano", _score_texto(criterios.get("tamano"), m.tamano, 3))
    _sumar("edad", _score_texto(criterios.get("edad"), m.edad, 2))
    _sumar("nombre", _score_texto(criterios.get("nombre"), m.nombre, 1))

    # "desconocido" es el valor por defecto de las plataformas de origen: que
    # dos reportes lo compartan no dice absolutamente nada, así que no puntúa.
    sexo = normalizar_sexo(criterios.get("sexo"))
    if sexo and sexo != "desconocido" and sexo == normalizar_sexo(m.sexo):
        _sumar("sexo", 2)

    # La zona SUMA cuando coincide, pero nunca descarta: un animal perdido en
    # San Fernando aparece a los tres días en Meléndez, y quien lo encontró
    # reporta dónde está, no dónde se perdió. Por eso pesa menos que el color o
    # la raza y jamás actúa como filtro.
    zona = criterios.get("zona")
    _sumar("zona", max(
        _score_texto(zona, m.ubicacion, 2, vacios=_ZONAS_GENERICAS),
        _score_texto(zona, m.barrio, 2, vacios=_ZONAS_GENERICAS),
    ))

    # Señas particulares / descripción libre contra todo el texto del reporte.
    descripcion = criterios.get("descripcion")
    if descripcion:
        blob = " ".join(
            filter(None, [m.senas, m.notas, m.color, m.raza, m.especie_otra])
        )
        _sumar("senas", min(5, len(_tokens(descripcion) & _tokens(blob))))

    return sum(detalle.values()), detalle


def _score_mascota(criterios: Dict[str, Any], m: models.Mascota) -> int:
    score, _ = _evaluar(criterios, m)
    return score


def ficha_publica(m: models.Mascota, db: Optional[Session] = None) -> Dict[str, Any]:
    """Datos de un reporte APTOS para mostrarle a un desconocido.

    Sin teléfono ni nombre de contacto: eso solo se entrega cuando la persona
    confirma que reconoce a la mascota (`datos_de_contacto`).
    """
    fotos = [f.id for f in (m.fotos or [])]
    externo = bool(m.origen_url)
    return {
        "codigo": m.codigo,
        # De dónde salió el reporte. El bot lo necesita para saber si el
        # contacto se entrega como teléfono (propio) o como link (importado).
        "externo": externo,
        "origen": ORIGEN_NOMBRES.get(m.source, m.source) if externo else None,
        "tipo": m.tipo_registro,
        "especie": m.especie_otra if m.especie == "otra" and m.especie_otra else m.especie,
        "raza": m.raza,
        "color": m.color,
        "nombre": m.nombre,
        "sexo": m.sexo,
        "edad": m.edad,
        "tamano": m.tamano,
        "senas": m.senas,
        "zona": m.barrio or m.ubicacion,
        "fecha": m.fecha_evento.isoformat() if m.fecha_evento else None,
        "fotos": len(fotos),
        # Las fotos de un reporte importado viven en el sitio de origen: no las
        # copiamos, así que su ficha se muestra con el enlace en vez de imagen.
        "foto_url": f"/mascotas/foto/{m.codigo}/{fotos[0]}" if fotos else None,
        "origen_url": m.origen_url,
        "reportado": m.created_at.strftime("%Y-%m-%d") if m.created_at else None,
    }


def buscar(
    db: Session,
    criterios: Dict[str, Any],
    buscar_en: str = "encontradas",
    limite: int = 4,
) -> List[Dict[str, Any]]:
    """Reportes que más se parecen a la descripción, mejor primero.

    `buscar_en`: 'encontradas' (alguien halló una mascota — el caso de quien
    busca a la suya), 'perdidas' (alguien la está buscando — el caso de quien
    encontró una) o 'todas'.
    """
    q = db.query(models.Mascota).filter(
        models.Mascota.estado == models.MASCOTA_ESTADO_ACTIVO
    )
    if buscar_en == "encontradas":
        q = q.filter(models.Mascota.tipo_registro == models.MASCOTA_TIPO_ENCONTRADA)
    elif buscar_en == "perdidas":
        q = q.filter(models.Mascota.tipo_registro == models.MASCOTA_TIPO_PERDIDA)

    especie = normalizar_especie(criterios.get("especie"))
    if especie:
        q = q.filter(models.Mascota.especie == especie)

    # El volumen esperado (reportes activos de una ciudad) cabe de sobra en
    # memoria; el scoring difuso no se puede expresar en SQL sin extensiones.
    candidatos = q.order_by(models.Mascota.created_at.desc()).limit(500).all()

    puntuados = []
    for m in candidatos:
        score = _score_mascota(criterios, m)
        if score >= 3:
            puntuados.append((score, m))
    puntuados.sort(key=lambda par: (-par[0], -(par[1].id or 0)))

    resultados = []
    for score, m in puntuados[:limite]:
        ficha = ficha_publica(m, db)
        ficha["coincidencia"] = score
        resultados.append(ficha)
    return resultados


# ---------------------------------------------------------------------------
# Alta y actualización de reportes
# ---------------------------------------------------------------------------

# Acepta uno o dos teléfonos separados por `/` o `,`: un animal en hogar de paso
# tiene dos vías de contacto (la fundación y la casa donde está durmiendo), y
# quien busca a su mascota necesita las dos.
_TELEFONO_RE = re.compile(r"^[+\d][\d\s\-()/,]{6,30}$")

# --- Quién está escribiendo, cuando todavía no hay reporte -------------------
# El panel mostraba el contacto de un hilo leyéndolo del reporte registrado. Si
# la persona daba su nombre y su teléfono en el primer mensaje y la conversación
# no llegaba a registrar nada, ese dato se perdía: el hilo quedaba anónimo y no
# había a quién devolverle la llamada. Estos dos extractores lo rescatan de lo
# que la persona escribió, sin esperar a que exista una ficha.

_TELEFONO_DICHO_RE = re.compile(r"(?:\+?57[\s-]?)?\b\d[\d\s\-()]{5,14}\d\b")

_NOMBRE_DICHO_RE = re.compile(
    r"\b(?:me\s+llamo|mi\s+nombre\s+es|habla(?:s)?\s+con|soy)\s+"
    r"([^\W\d_]{2,}(?:\s+[^\W\d_]{2,})?)",
    re.IGNORECASE | re.UNICODE,
)

# Palabras que siguen a "soy" sin ser un nombre: "soy de Cali", "soy la dueña".
# Sin esta lista el panel se llena de contactos llamados "De" o "La".
_NO_ES_NOMBRE = frozenset({
    "de", "del", "la", "el", "los", "las", "un", "una", "unos", "unas",
    "mi", "su", "tu", "yo", "quien", "que", "muy", "solo", "solamente",
    "dueno", "duena", "dueño", "dueña", "papa", "mama", "papá", "mamá",
    "hermano", "hermana", "vecino", "vecina", "amigo", "amiga", "hijo",
    "hija", "persona", "usuario", "nuevo", "nueva", "aqui", "aquí",
})


def _sin_tildes(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def telefono_dicho(texto: str) -> Optional[str]:
    """El primer teléfono que aparece en lo que escribió la persona."""
    for candidato in _TELEFONO_DICHO_RE.findall(texto or ""):
        digitos = re.sub(r"\D", "", candidato)
        # 7 dígitos = fijo sin indicativo. Por debajo son fechas, horas o
        # códigos de reporte; por encima de 15 no es un número marcable.
        if 7 <= len(digitos) <= 15:
            return digitos
    return None


def nombre_dicho(texto: str) -> Optional[str]:
    """El nombre con el que la persona se presentó, si se presentó."""
    m = _NOMBRE_DICHO_RE.search(texto or "")
    if m is None:
        return None
    partes = [p for p in m.group(1).split() if p]
    if not partes or _sin_tildes(partes[0]).lower() in _NO_ES_NOMBRE:
        return None
    return " ".join(p.capitalize() for p in partes)[:120]


def contacto_dicho(texto: str) -> Optional[str]:
    """Cómo identificar a quien escribe: "Ana · 3001234567", o lo que haya."""
    nombre, telefono = nombre_dicho(texto), telefono_dicho(texto)
    if nombre and telefono:
        return f"{nombre} · {telefono}"
    return nombre or telefono


def _limpiar(valor: Any, limite: int) -> Optional[str]:
    if valor is None:
        return None
    texto = re.sub(r"\s+", " ", str(valor).replace("\x00", "")).strip()
    return texto[:limite] or None


def _parse_fecha(valor: Any) -> Optional[date]:
    texto = _limpiar(valor, 10)
    if not texto:
        return None
    try:
        return date.fromisoformat(texto)
    except ValueError:
        return None


def _parse_datetime(valor: Any) -> Optional[datetime]:
    """ISO 8601 con o sin zona. Las fuentes lo mandan de las dos formas."""
    if isinstance(valor, datetime):
        return valor
    texto = _limpiar(valor, 40)
    if not texto:
        return None
    try:
        # `fromisoformat` de 3.11 acepta la Z, pero guardamos naive en UTC
        # para que sea consistente con el resto de las columnas de la tabla.
        parsed = datetime.fromisoformat(texto.replace("Z", "+00:00"))
        return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
    except ValueError:
        return None


def _booleano(valor: Any) -> Optional[bool]:
    """Tri-estado: None cuando la fuente no dice nada.

    Es la diferencia entre "no sabemos si está esterilizada" y "nos dijeron que
    no lo está", y en una ficha de adopción esa diferencia importa.
    """
    if valor is None or valor == "":
        return None
    if isinstance(valor, bool):
        return valor
    texto = normalizar(str(valor))
    if texto in ("si", "s", "true", "1", "yes", "y"):
        return True
    if texto in ("no", "n", "false", "0"):
        return False
    return None          # "NO SE SABE", "AUN NO SE SABE", cualquier otra cosa


def _decimal(valor: Any) -> Optional[float]:
    try:
        numero = float(str(valor).replace(",", "."))
    except (TypeError, ValueError):
        return None
    # Las fuentes mandan 0 cuando el campo está sin usar (Protección Animal
    # tiene peso_animal=0 en todas sus fichas), y un animal de 0 kg no existe.
    return round(numero, 2) if 0 < numero < 999 else None


def crear_reporte(
    db: Session,
    datos: Dict[str, Any],
    *,
    bot_id: Optional[int] = None,
    source: str = "web",
    upload_session: Optional[str] = None,
) -> Tuple[Optional[models.Mascota], str]:
    """Crea un reporte. Devuelve (mascota, problema).

    `problema` es un mensaje para el MODELO (no para el ciudadano): explica qué
    dato falta o vino mal para que el bot lo pida de nuevo y reintente.
    """
    tipo = normalizar(datos.get("tipo_registro"))
    if tipo not in models.AVAILABLE_MASCOTA_TIPOS:
        return None, (
            "el tipo de reporte debe ser 'perdida' (la persona busca a su "
            "mascota) o 'encontrada' (la persona halló una mascota)"
        )

    especie = normalizar_especie(datos.get("especie"))
    if not especie:
        return None, "falta saber si es perro, gato u otra especie: pregúntaselo"

    ubicacion = _limpiar(datos.get("ubicacion"), 255)
    if not ubicacion:
        return None, (
            "falta la ubicación (dónde se perdió o dónde fue vista): es "
            "obligatoria, pídesela con un ejemplo del tipo 'barrio y calle'"
        )

    telefono = _limpiar(datos.get("contacto_telefono"), 32)
    origen_url = _limpiar(datos.get("origen_url"), 500)
    if not telefono and not origen_url:
        # Un reporte necesita una vía de contacto. Los que entran por el bot
        # traen teléfono; los importados de otra plataforma no lo publican y
        # se resuelven mandando a su ficha original.
        return None, (
            "falta un teléfono de contacto válido (con indicativo si es fijo): "
            "pídeselo, sin eso nadie puede avisarle"
        )
    if telefono and not _TELEFONO_RE.match(telefono):
        return None, (
            "ese teléfono no parece válido: pídeselo de nuevo con amabilidad"
        )

    mascota = models.Mascota(
        codigo="PENDIENTE",
        tipo_registro=tipo,
        especie=especie,
        especie_otra=_limpiar(datos.get("especie_otra"), 60) if especie == "otra" else None,
        raza=_limpiar(datos.get("raza"), 80),
        color=_limpiar(datos.get("color"), 80),
        nombre=_limpiar(datos.get("nombre"), 80),
        sexo=normalizar_sexo(datos.get("sexo")),
        edad=_limpiar(datos.get("edad"), 40),
        tamano=_limpiar(datos.get("tamano"), 24),
        senas=_limpiar(datos.get("senas"), 2000),
        ubicacion=ubicacion,
        maps_url=_limpiar(datos.get("maps_url"), 500),
        barrio=_limpiar(datos.get("barrio"), 120),
        contacto_nombre=_limpiar(datos.get("contacto_nombre"), 120),
        contacto_telefono=telefono,
        fecha_evento=_parse_fecha(datos.get("fecha_evento")),
        notas=_limpiar(datos.get("notas"), 2000),
        origen_url=origen_url,
        origen_id=_limpiar(datos.get("origen_id"), 120),
        bot_id=bot_id,
        source=source[:24],
        estado=models.MASCOTA_ESTADO_ACTIVO,
        # Campos multi-fuente: los llenan los importadores con lo que publica
        # cada plataforma. El bot no los pide — nadie que esté buscando a su
        # perro sabe si lo desparasitaron.
        ciudad=_limpiar(datos.get("ciudad"), 120),
        departamento=_limpiar(datos.get("departamento"), 120),
        esterilizado=_booleano(datos.get("esterilizado")),
        vacunado=_booleano(datos.get("vacunado")),
        desparasitado=_booleano(datos.get("desparasitado")),
        peso_kg=_decimal(datos.get("peso_kg")),
        salud=_limpiar(datos.get("salud"), 255),
        resguardo=_limpiar(datos.get("resguardo"), 40),
        resguardo_nombre=_limpiar(datos.get("resguardo_nombre"), 120),
        rescatado_por=_limpiar(datos.get("rescatado_por"), 120),
        rescatado_por_telefono=_limpiar(datos.get("rescatado_por_telefono"), 32),
        recompensa=_booleano(datos.get("recompensa")),
        estado_origen=_limpiar(datos.get("estado_origen"), 60),
        publicado_origen_at=_parse_datetime(datos.get("publicado_origen_at")),
        sincronizado_at=datetime.utcnow() if origen_url else None,
    )
    db.add(mascota)
    db.flush()   # necesitamos el id para derivar el código de la carpeta
    mascota.codigo = f"MC-{mascota.id:05d}"
    db.commit()
    db.refresh(mascota)

    if upload_session:
        adoptar_fotos(db, mascota, upload_session)
        db.refresh(mascota)

    # Sin PII en el log (regla #1): solo el código y el tipo.
    logger.info(
        "mascota reporte creado codigo=%s tipo=%s especie=%s source=%s",
        mascota.codigo, mascota.tipo_registro, mascota.especie, mascota.source,
    )
    return mascota, ""


_CAMPOS_ACTUALIZABLES = (
    "raza", "color", "nombre", "sexo", "edad", "tamano", "senas", "ubicacion",
    "maps_url", "barrio", "contacto_nombre", "contacto_telefono", "notas",
)
_LIMITES = {
    "raza": 80, "color": 80, "nombre": 80, "sexo": 16, "edad": 40, "tamano": 24,
    "senas": 2000, "ubicacion": 255, "maps_url": 500, "barrio": 120,
    "contacto_nombre": 120, "contacto_telefono": 32, "notas": 2000,
    # Solo editables desde el panel, no por el bot.
    "especie_otra": 60, "especie": 24, "tipo_registro": 16, "estado": 24,
    # Multi-fuente: los llenan los importadores y el equipo los corrige a mano.
    "ciudad": 120, "departamento": 120, "salud": 255, "resguardo": 40,
    "resguardo_nombre": 120, "rescatado_por": 120, "rescatado_por_telefono": 32,
    "estado_origen": 60,
}
# Los tri-estado del panel van por su propio camino: `_limpiar` los volvería
# la cadena "False", que en SQL es tan verdadera como "True".
_CAMPOS_BOOLEANOS = ("esterilizado", "vacunado", "desparasitado", "recompensa")


def actualizar_reporte(
    db: Session, codigo: str, datos: Dict[str, Any]
) -> Tuple[Optional[models.Mascota], str]:
    """Completa datos de un reporte ya creado (la gente los manda de a poco)."""
    mascota = (
        db.query(models.Mascota)
        .filter(models.Mascota.codigo == (codigo or "").strip().upper())
        .first()
    )
    if mascota is None:
        return None, f"no existe ningún reporte con el código {codigo}"

    cambios = []
    for campo in _CAMPOS_ACTUALIZABLES:
        if campo not in datos or datos[campo] in (None, ""):
            continue
        valor = _limpiar(datos[campo], _LIMITES[campo])
        if campo == "sexo":
            valor = normalizar_sexo(valor)
        if valor:
            setattr(mascota, campo, valor)
            cambios.append(campo)

    fecha = _parse_fecha(datos.get("fecha_evento"))
    if fecha:
        mascota.fecha_evento = fecha
        cambios.append("fecha_evento")

    if not cambios:
        return mascota, "no había datos nuevos que guardar"
    mascota.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(mascota)
    logger.info("mascota reporte actualizado codigo=%s campos=%s",
                mascota.codigo, ",".join(cambios))
    return mascota, ""


# Campos que el equipo puede corregir desde el panel. Incluye los dos que el
# bot exige al crear (`ubicacion` y `contacto_telefono`): se pueden corregir,
# nunca dejar vacíos — un reporte sin dónde ni a quién llamar no sirve para
# reunir a nadie.
_CAMPOS_PANEL = (
    "tipo_registro", "especie", "especie_otra", "raza", "color", "nombre",
    "sexo", "edad", "tamano", "senas", "ubicacion", "maps_url", "barrio",
    "contacto_nombre", "contacto_telefono", "fecha_evento", "estado", "notas",
    # Multi-fuente. `peso_kg` y las fechas de origen no se editan a mano: son
    # datos de la fuente, y corregirlos a dedo rompería la trazabilidad.
    "ciudad", "departamento", "salud", "resguardo", "resguardo_nombre",
    "rescatado_por", "rescatado_por_telefono", "estado_origen",
    "esterilizado", "vacunado", "desparasitado", "recompensa",
)


def editar_desde_panel(
    db: Session, codigo: str, datos: Dict[str, Any]
) -> Tuple[Optional[models.Mascota], str]:
    """Edición manual de un reporte. Devuelve (mascota, problema).

    A diferencia de `actualizar_reporte` (que usa el bot y solo agrega datos),
    aquí se puede **corregir y vaciar** campos opcionales, y cambiar el tipo o
    el estado. Los obligatorios se validan explícitamente.
    """
    mascota = obtener(db, codigo)
    if mascota is None:
        return None, "El reporte no existe"

    cambios = []
    for campo in _CAMPOS_PANEL:
        if campo not in datos:
            continue

        if campo == "fecha_evento":
            valor = _parse_fecha(datos[campo])
            if datos[campo] and valor is None:
                return None, "La fecha debe tener el formato AAAA-MM-DD"
        elif campo == "tipo_registro":
            valor = normalizar(datos[campo])
            if valor not in models.AVAILABLE_MASCOTA_TIPOS:
                return None, "El tipo de reporte debe ser 'perdida' o 'encontrada'"
        elif campo == "estado":
            valor = normalizar(datos[campo])
            if valor not in models.AVAILABLE_MASCOTA_ESTADOS:
                return None, "Estado inválido"
        elif campo == "especie":
            valor = normalizar_especie(datos[campo])
            if not valor:
                return None, "La especie es obligatoria (perro, gato u otra)"
        elif campo == "sexo":
            valor = normalizar_sexo(datos[campo])
        elif campo in _CAMPOS_BOOLEANOS:
            # Se puede volver a "no sabemos" mandando null o cadena vacía: que
            # el equipo pueda deshacer un dato mal puesto importa más que
            # forzarlo a elegir entre sí y no.
            valor = _booleano(datos[campo])
        else:
            valor = _limpiar(datos[campo], _LIMITES.get(campo, 255))

        if campo == "ubicacion" and not valor:
            return None, (
                "La ubicación es obligatoria: dónde se perdió, dónde se "
                "encontró o dónde está ahora la mascota"
            )
        if campo == "contacto_telefono":
            if not valor:
                return None, "El teléfono de contacto es obligatorio"
            if not _TELEFONO_RE.match(valor):
                return None, "El teléfono no parece válido"

        if getattr(mascota, campo) != valor:
            setattr(mascota, campo, valor)
            cambios.append(campo)

    if not cambios:
        return mascota, ""
    mascota.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(mascota)
    # Sin PII en el log (regla #1): solo el código y qué campos cambiaron.
    logger.info("mascotas panel: %s editado campos=%s", mascota.codigo, ",".join(cambios))
    return mascota, ""


def eliminar_reporte(db: Session, codigo: str) -> bool:
    """Borra un reporte, sus fotos (fila y archivo) y sus coincidencias.

    Las coincidencias y las filas de fotos caen por `ON DELETE CASCADE`; los
    archivos del storage hay que borrarlos a mano o quedan huérfanos pagando
    almacenamiento.
    """
    mascota = obtener(db, codigo)
    if mascota is None:
        return False
    for foto in list(mascota.fotos or []):
        _delete_object(foto.storage_key)
    db.delete(mascota)
    db.commit()
    logger.info("mascotas panel: reporte %s eliminado", codigo)
    return True


def eliminar_foto(db: Session, codigo: str, foto_id: int) -> bool:
    """Borra una sola foto de un reporte (la fila y el archivo)."""
    foto = (
        db.query(models.MascotaFoto)
        .join(models.Mascota, models.MascotaFoto.mascota_id == models.Mascota.id)
        .filter(
            models.MascotaFoto.id == foto_id,
            models.Mascota.codigo == (codigo or "").strip().upper(),
        )
        .first()
    )
    if foto is None:
        return False
    _delete_object(foto.storage_key)
    db.delete(foto)
    db.commit()
    logger.info("mascotas panel: foto %s de %s eliminada", foto_id, codigo)
    return True


def purgar(db: Session, source: str) -> int:
    """Borra de un golpe todos los reportes de un origen.

    Pensado para dejar la base limpia de datos de prueba (`source='demo'`)
    antes de abrir al público, sin tocar los reportes reales.
    """
    reportes = (
        db.query(models.Mascota).filter(models.Mascota.source == source).all()
    )
    for mascota in reportes:
        for foto in list(mascota.fotos or []):
            _delete_object(foto.storage_key)
        db.delete(mascota)
    if reportes:
        db.commit()
    logger.info("mascotas panel: %s reportes de origen %r eliminados",
                len(reportes), source)
    return len(reportes)


def _reconocer_una(mascota, chat_ref: Optional[str]) -> bool:
    """Pasa UNA ficha a `reconocida`. False si su estado ya no lo permite."""
    if mascota is None or mascota.estado != models.MASCOTA_ESTADO_ACTIVO:
        return False
    mascota.estado = models.MASCOTA_ESTADO_RECONOCIDA
    mascota.reconocida_at = datetime.utcnow()
    mascota.reconocida_chat = (chat_ref or "")[:64] or None
    mascota.updated_at = datetime.utcnow()
    return True


def marcar_reconocida(
    db: Session,
    codigo: str,
    chat_ref: Optional[str] = None,
    codigo_perdida: Optional[str] = None,
) -> bool:
    """Deja constancia de que alguien dijo reconocer a esta mascota.

    Es una afirmación sin verificar: pasa a `reconocida` ("por confirmar"), no a
    `reunida`. El equipo llama a las dos partes y recién entonces confirma el
    reencuentro desde el panel.

    Un reencuentro tiene DOS fichas y un par en `mascota_coincidencias`, y antes
    solo se marcaba la encontrada. Las otras dos filas se quedaban como si nada
    hubiera pasado: la familia seguía apareciendo "buscando" y la coincidencia
    seguía contando como "sin revisar", que es justo la lista por la que el
    equipo decide a quién llamar. Cuando la conversación registró el reporte de
    quien busca (`codigo_perdida`), se marcan los tres.

    No toca los casos cerrados, reunidos o ya reconocidos: el estado solo
    avanza, nunca retrocede por una conversación nueva.
    """
    encontrada = obtener(db, codigo)
    marcada = _reconocer_una(encontrada, chat_ref)

    perdida = obtener(db, codigo_perdida) if codigo_perdida else None
    if perdida is not None and perdida.id != getattr(encontrada, "id", None):
        if _reconocer_una(perdida, chat_ref):
            logger.info(
                "mascotas: %s (quien busca) marcada como reconocida", perdida.codigo
            )

    # El par del panel. Solo el que une a estas dos fichas: la encontrada puede
    # tener varias coincidencias abiertas y solo una persona la reconoció.
    if encontrada is not None and perdida is not None:
        par = (
            db.query(models.MascotaCoincidencia)
            .filter(
                models.MascotaCoincidencia.perdida_id == perdida.id,
                models.MascotaCoincidencia.encontrada_id == encontrada.id,
                models.MascotaCoincidencia.estado.in_(
                    (models.MATCH_ESTADO_NUEVA, models.MATCH_ESTADO_REVISADA)
                ),
            )
            .first()
        )
        if par is not None:
            par.estado = models.MATCH_ESTADO_RECONOCIDA
            par.updated_at = datetime.utcnow()
            logger.info(
                "mascotas: coincidencia %s↔%s marcada como reconocida",
                perdida.codigo, encontrada.codigo,
            )

    db.commit()
    if marcada:
        logger.info(
            "mascotas: %s marcada como reconocida (por confirmar)", encontrada.codigo
        )
    return marcada


def datos_de_contacto(db: Session, codigo: str) -> Optional[Dict[str, Any]]:
    """Ubicación exacta + teléfono. Solo cuando la persona reconoce a la
    mascota — es el único momento en que se entrega PII de quien reportó."""
    mascota = (
        db.query(models.Mascota)
        .filter(models.Mascota.codigo == (codigo or "").strip().upper())
        .first()
    )
    if mascota is None:
        return None
    datos = {
        "codigo": mascota.codigo,
        "tipo": mascota.tipo_registro,
        "ubicacion": mascota.ubicacion,
        "maps_url": mascota.maps_url,
        "barrio": mascota.barrio,
        "contacto_nombre": mascota.contacto_nombre,
        "contacto_telefono": mascota.contacto_telefono,
    }
    if mascota.origen_url:
        # Reporte importado de otra plataforma: no tenemos su teléfono, y el
        # contacto se resuelve mandando a la ficha original.
        datos["origen_url"] = mascota.origen_url
        datos["origen"] = ORIGEN_NOMBRES.get(mascota.source, mascota.source)
    return datos


def obtener(db: Session, codigo: str) -> Optional[models.Mascota]:
    return (
        db.query(models.Mascota)
        .filter(models.Mascota.codigo == (codigo or "").strip().upper())
        .first()
    )


# ---------------------------------------------------------------------------
# Cruce diario perdidas ↔ encontradas
# ---------------------------------------------------------------------------

# Umbral del job. Mucho más alto que el de la búsqueda en vivo (3) a propósito:
# aquí nadie confirma datos al otro lado del chat, así que una coincidencia
# floja solo genera trabajo inútil. Con ~250 perdidas y ~50 encontradas hay
# 12.500 pares posibles; con umbral 6 pasaban 5.284 (todo perro negro parecido
# a todo perro negro) y el panel se volvía inservible. Con 12 solo pasan los
# pares que comparten varias señas concretas.
UMBRAL_COINCIDENCIA = 12

# Cuántas candidatas se guardan por cada mascota buscada. El equipo llama de a
# una: más de tres por caso es ruido, y las mejores están siempre arriba.
MAX_COINCIDENCIAS_POR_PERDIDA = 3


def _criterios_de(m: models.Mascota) -> Dict[str, Any]:
    """Convierte un reporte en los criterios de búsqueda que lo describen."""
    return {
        "especie": m.especie,
        "raza": m.raza,
        "color": m.color,
        "nombre": m.nombre,
        "sexo": m.sexo,
        "edad": m.edad,
        "tamano": m.tamano,
        "zona": " ".join(filter(None, [m.barrio, m.ubicacion])),
        "descripcion": " ".join(filter(None, [m.senas, m.notas])),
    }


def cruzar_reportes(db: Session, umbral: int = UMBRAL_COINCIDENCIA) -> Dict[str, int]:
    """Compara cada 'perdida' activa contra cada 'encontrada' activa.

    Lo corre el job diario de las 12:00 (hora Colombia). Es idempotente: si un
    par ya estaba registrado actualiza su puntaje y **respeta el estado** que le
    haya puesto el equipo (una coincidencia descartada no vuelve a "nueva").

    Devuelve contadores para el log del job.
    """
    perdidas = (
        db.query(models.Mascota)
        .filter(
            models.Mascota.tipo_registro == models.MASCOTA_TIPO_PERDIDA,
            models.Mascota.estado == models.MASCOTA_ESTADO_ACTIVO,
        )
        .all()
    )
    encontradas = (
        db.query(models.Mascota)
        .filter(
            models.Mascota.tipo_registro == models.MASCOTA_TIPO_ENCONTRADA,
            models.Mascota.estado == models.MASCOTA_ESTADO_ACTIVO,
        )
        .all()
    )

    existentes = {
        (c.perdida_id, c.encontrada_id): c
        for c in db.query(models.MascotaCoincidencia).all()
    }

    nuevas = actualizadas = evaluados = 0
    for perdida in perdidas:
        criterios = _criterios_de(perdida)

        # Se evalúan todas las candidatas y solo se guardan las mejores: sin
        # este recorte, una descripción genérica ("perro negro") generaba
        # decenas de pares por caso y enterraba a los que sí valen la pena.
        candidatas = []
        for encontrada in encontradas:
            evaluados += 1
            score, detalle = _evaluar(criterios, encontrada)
            if score >= umbral:
                candidatas.append((score, encontrada, detalle))
        candidatas.sort(key=lambda c: (-c[0], -(c[1].id or 0)))

        for score, encontrada, detalle in candidatas[:MAX_COINCIDENCIAS_POR_PERDIDA]:
            clave = (perdida.id, encontrada.id)
            existente = existentes.get(clave)
            if existente is None:
                db.add(models.MascotaCoincidencia(
                    perdida_id=perdida.id,
                    encontrada_id=encontrada.id,
                    score=score,
                    detalle=detalle,
                    estado=models.MATCH_ESTADO_NUEVA,
                ))
                nuevas += 1
            elif existente.score != score or existente.detalle != detalle:
                existente.score = score
                existente.detalle = detalle
                existente.updated_at = datetime.utcnow()
                actualizadas += 1
    db.commit()

    logger.info(
        "mascotas cruce: perdidas=%s encontradas=%s pares=%s nuevas=%s actualizadas=%s",
        len(perdidas), len(encontradas), evaluados, nuevas, actualizadas,
    )
    return {
        "perdidas": len(perdidas),
        "encontradas": len(encontradas),
        "pares_evaluados": evaluados,
        "nuevas": nuevas,
        "actualizadas": actualizadas,
    }


def listar_coincidencias(
    db: Session, estado: Optional[str] = None, limite: int = 300
) -> List[models.MascotaCoincidencia]:
    q = db.query(models.MascotaCoincidencia)
    if estado in models.AVAILABLE_MATCH_ESTADOS:
        q = q.filter(models.MascotaCoincidencia.estado == estado)
    return (
        q.order_by(
            models.MascotaCoincidencia.score.desc(),
            models.MascotaCoincidencia.created_at.desc(),
        )
        .limit(limite)
        .all()
    )


# ---------------------------------------------------------------------------
# Listado y exportación
# ---------------------------------------------------------------------------

COLUMNAS_EXCEL = [
    ("codigo", "Código"),
    ("tipo_registro", "Tipo de reporte"),
    ("especie", "Especie"),
    ("raza", "Raza"),
    ("color", "Color"),
    ("nombre", "Nombre"),
    ("sexo", "Sexo"),
    ("edad", "Edad"),
    ("tamano", "Tamaño"),
    ("senas", "Señas particulares"),
    ("notas", "Comentarios adicionales"),
    ("ubicacion", "Ubicación"),
    ("barrio", "Barrio / zona"),
    ("maps_url", "Google Maps"),
    ("fecha_evento", "Fecha del hecho"),
    ("contacto_nombre", "Contacto"),
    ("contacto_telefono", "Teléfono"),
    ("estado", "Estado"),
    ("fotos", "Fotos"),
    ("origen", "Origen"),
    ("origen_url", "Ficha original"),
    ("created_at", "Reportado el"),
]


def listar(
    db: Session,
    tipo: Optional[str] = None,
    estado: Optional[str] = None,
    limite: int = 1000,
) -> List[models.Mascota]:
    q = db.query(models.Mascota)
    if tipo in models.AVAILABLE_MASCOTA_TIPOS:
        q = q.filter(models.Mascota.tipo_registro == tipo)
    if estado in models.AVAILABLE_MASCOTA_ESTADOS:
        q = q.filter(models.Mascota.estado == estado)
    return q.order_by(models.Mascota.created_at.desc()).limit(limite).all()


def _valor_excel(m: models.Mascota, campo: str) -> Any:
    if campo == "fotos":
        return len(m.fotos or [])
    if campo == "origen":
        return ORIGEN_NOMBRES.get(m.source, "Recupera Tu Mascota")
    if campo == "tipo_registro":
        return "Mascota encontrada" if m.tipo_registro == "encontrada" else "Mascota perdida"
    if campo == "especie":
        return m.especie_otra if m.especie == "otra" and m.especie_otra else m.especie
    valor = getattr(m, campo, None)
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d %H:%M")
    if isinstance(valor, date):
        return valor.isoformat()
    return valor


def exportar_excel(db: Session, tipo: Optional[str] = None) -> bytes:
    """Listado vigente en .xlsx (una hoja, encabezado congelado y autofiltro)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    registros = listar(db, tipo=tipo, estado=models.MASCOTA_ESTADO_ACTIVO)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mascotas"

    encabezado_fill = PatternFill("solid", fgColor="004D40")
    encabezado_font = Font(color="FFFFFF", bold=True)
    for col, (_, titulo) in enumerate(COLUMNAS_EXCEL, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(vertical="center")

    for fila, m in enumerate(registros, start=2):
        for col, (campo, _) in enumerate(COLUMNAS_EXCEL, start=1):
            ws.cell(row=fila, column=col, value=_valor_excel(m, campo))

    anchos = {"senas": 40, "ubicacion": 36, "maps_url": 30, "raza": 18, "color": 18}
    for col, (campo, titulo) in enumerate(COLUMNAS_EXCEL, start=1):
        ws.column_dimensions[get_column_letter(col)].width = anchos.get(
            campo, max(12, min(24, len(titulo) + 4))
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLUMNAS_EXCEL))}{max(1, len(registros) + 1)}"
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Export para plataformas aliadas
# ---------------------------------------------------------------------------

# Versión del formato de intercambio. Si cambia la forma de los registros, sube
# el número: quien consuma el archivo puede así saber qué esperar.
EXPORT_VERSION = "1.0"


def exportar_json(db: Session, incluir_contacto: bool = False) -> Dict[str, Any]:
    """Todos los casos en un JSON pensado para compartir con otra plataforma.

    Por defecto **sin datos de contacto**: los teléfonos son de ciudadanos que
    los dieron para que los llamen por su mascota, no para redistribuirlos. Cada
    registro lleva `origen_url` (si vino de otra plataforma) o el enlace a
    nuestro chat, para que quien reconozca al animal sepa dónde escribir.
    """
    registros = []
    for m in listar(db, limite=100000):
        registro = {
            "codigo": m.codigo,
            "tipo": m.tipo_registro,
            "estado": m.estado,
            "especie": m.especie,
            "especie_otra": m.especie_otra,
            "raza": m.raza,
            "color": m.color,
            "nombre": m.nombre,
            "sexo": m.sexo,
            "edad": m.edad,
            "tamano": m.tamano,
            "senas": m.senas,
            "ubicacion": m.ubicacion,
            "barrio": m.barrio,
            "maps_url": m.maps_url,
            "fecha_evento": m.fecha_evento.isoformat() if m.fecha_evento else None,
            "reportado_el": m.created_at.isoformat() if m.created_at else None,
            "origen": ORIGEN_NOMBRES.get(m.source, "Recupera Tu Mascota"),
            "origen_url": m.origen_url,
            "fotos": [f"fotos/{m.codigo}/{f.id}.jpg" for f in (m.fotos or [])],
            "contacto_url": m.origen_url or "https://mascotasperdidascolombia.com",
        }
        if incluir_contacto:
            registro["contacto_nombre"] = m.contacto_nombre
            registro["contacto_telefono"] = m.contacto_telefono
        registros.append(registro)

    return {
        "formato": "recupera-tu-mascota/export",
        "version": EXPORT_VERSION,
        "generado": datetime.utcnow().isoformat() + "Z",
        "fuente": "https://mascotasperdidascolombia.com",
        "licencia": (
            "Datos compartidos para reunir mascotas perdidas con sus familias. "
            "No usar con fines comerciales ni de contacto masivo."
        ),
        "contacto_incluido": incluir_contacto,
        "total": len(registros),
        "casos": registros,
    }


def exportar_zip(db: Session, incluir_contacto: bool = False,
                 incluir_fotos: bool = True) -> bytes:
    """Paquete para la app amiga: el JSON de casos + las fotos que tengamos.

    Estructura del archivo:
        casos.json                 → todos los registros
        LEEME.txt                  → qué es esto y cómo leerlo
        fotos/<codigo>/<id>.jpg    → las fotos de cada caso

    Solo se empaquetan las fotos que viven en NUESTRO storage. Las de reportes
    importados no están: viven en la plataforma de origen y el JSON lleva su
    `origen_url`.
    """
    import zipfile

    datos = exportar_json(db, incluir_contacto=incluir_contacto)
    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "casos.json", json.dumps(datos, ensure_ascii=False, indent=1)
        )
        zf.writestr("LEEME.txt", _texto_leeme(datos, incluir_fotos))

        if incluir_fotos:
            for m in listar(db, limite=100000):
                for foto in (m.fotos or []):
                    contenido = _get_object(foto.storage_key)
                    if contenido is None:
                        continue
                    extension = ALLOWED_IMAGE_TYPES.get(
                        foto.content_type or "image/jpeg", ".jpg"
                    )
                    zf.writestr(f"fotos/{m.codigo}/{foto.id}{extension}", contenido)

    return buffer.getvalue()


def _texto_leeme(datos: Dict[str, Any], incluir_fotos: bool) -> str:
    return "\n".join([
        "Recupera Tu Mascota — export de casos",
        "=" * 38,
        "",
        f"Generado: {datos['generado']}",
        f"Casos: {datos['total']}",
        f"Formato: {datos['formato']} v{datos['version']}",
        "",
        "Contenido",
        "---------",
        "casos.json  Todos los reportes: mascotas perdidas (las busca su",
        "            familia) y encontradas (alguien las halló). El campo",
        "            'tipo' distingue unas de otras.",
        "fotos/      Fotos de cada caso, en una carpeta por código." if incluir_fotos
        else "            (Este paquete se generó sin fotos.)",
        "",
        "Datos de contacto",
        "-----------------",
        ("Este paquete INCLUYE teléfonos de contacto. Trátalos como datos "
         "personales: son de ciudadanos que los dieron para que los llamen por "
         "su mascota."
         if datos["contacto_incluido"] else
         "Este paquete NO incluye teléfonos. Cada caso trae 'contacto_url' con "
         "el sitio donde escribir para llegar a quien reportó."),
        "",
        "Casos importados",
        "----------------",
        "Los que traen 'origen_url' vienen de otra plataforma solidaria; ese",
        "enlace es la ficha original, con sus propios datos de contacto.",
        "",
        datos["licencia"],
        "",
        f"Fuente: {datos['fuente']}",
    ])


def resumen(db: Session) -> Dict[str, Any]:
    """Contadores del panel: totales por tipo, especie y estado."""
    por_tipo = dict(
        db.query(models.Mascota.tipo_registro, func.count(models.Mascota.id))
        .group_by(models.Mascota.tipo_registro)
        .all()
    )
    por_especie = dict(
        db.query(models.Mascota.especie, func.count(models.Mascota.id))
        .filter(models.Mascota.estado == models.MASCOTA_ESTADO_ACTIVO)
        .group_by(models.Mascota.especie)
        .all()
    )
    por_estado = dict(
        db.query(models.Mascota.estado, func.count(models.Mascota.id))
        .group_by(models.Mascota.estado)
        .all()
    )
    total_fotos = db.query(func.count(models.MascotaFoto.id)).filter(
        models.MascotaFoto.mascota_id.isnot(None)
    ).scalar()
    return {
        "total": sum(por_tipo.values()),
        "perdidas": por_tipo.get(models.MASCOTA_TIPO_PERDIDA, 0),
        "encontradas": por_tipo.get(models.MASCOTA_TIPO_ENCONTRADA, 0),
        "activas": por_estado.get(models.MASCOTA_ESTADO_ACTIVO, 0),
        "reconocidas": por_estado.get(models.MASCOTA_ESTADO_RECONOCIDA, 0),
        "reunidas": por_estado.get(models.MASCOTA_ESTADO_REUNIDA, 0),
        "cerradas": por_estado.get(models.MASCOTA_ESTADO_CERRADO, 0),
        "por_especie": por_especie,
        "fotos": int(total_fotos or 0),
    }
