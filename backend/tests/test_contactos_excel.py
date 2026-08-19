"""Importación de contactos por Excel.

Lo que se prueba aquí es el contrato que ve una persona **no técnica**: sube el
archivo que descargó de la plataforma y espera un resumen en español, no un
error de validación. Por eso las tres pruebas que exige el módulo son:

1. Un Excel bien llenado entra completo, incluidas las columnas extra como
   atributos del contacto.
2. Un Excel con teléfonos malos **no tumba la importación**: las filas buenas
   entran y cada rechazo dice fila + motivo, sin el número (regla 1).
3. Reimportar el mismo archivo no duplica: la segunda pasada es toda
   "actualizados" y la cuenta de contactos no se mueve.

Todo corre contra SQLite en memoria, sin Postgres ni red (igual que el resto
de la suite; el CHECK del teléfono lo quita `backend/conftest.py` porque usa
el operador `~` de Postgres, y la validación equivalente vive en el importador).
"""
from __future__ import annotations

import io

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app import models
from app.services import contactos_excel


# ---------------------------------------------------------------------------
# Infra
# ---------------------------------------------------------------------------

@pytest.fixture
def Sesion():
    from app.database import Base

    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    yield sessionmaker(autocommit=False, autoflush=False, bind=engine)
    engine.dispose()


@pytest.fixture
def db(Sesion):
    sesion = Sesion()
    try:
        yield sesion
    finally:
        sesion.close()


@pytest.fixture
def team(db):
    """Un team mínimo. `Contact.team_id` es FK, así que tiene que existir."""
    usuario = models.User(
        nombre="Agencia Demo",
        tipo_documento="NIT",
        documento="900123456",
        correo="agencia@demo.com",
        hashed_password="x",
    )
    db.add(usuario)
    db.commit()
    equipo = models.Team(nombre="Agencia Demo", owner_user_id=usuario.id)
    db.add(equipo)
    db.commit()
    return equipo


def hacer_excel(filas: list[list], hoja: str = contactos_excel.HOJA_DATOS) -> bytes:
    """Arma un .xlsx en memoria. `filas[0]` son los encabezados."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    for fila in filas:
        ws.append(fila)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


ENCABEZADOS = ["Nombre", "Teléfono", "Correo", "Acepta mensajes", "Ciudad", "Idioma"]


# ---------------------------------------------------------------------------
# 1. Excel válido
# ---------------------------------------------------------------------------

class TestExcelValido:
    def test_crea_los_contactos_y_reporta_el_resumen(self, db, team):
        contenido = hacer_excel([
            ENCABEZADOS,
            ["Ana Ruiz", "+57 300 111 2233", "ana@ejemplo.com", "sí", "Cali", "Español"],
            ["Luis Paz", "+52 55 1234 5678", "", "sí", "CDMX", "Español"],
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert (res.total, res.created, res.updated, res.rejected) == (2, 2, 0, 0)
        assert res.errors == []
        assert db.query(models.Contact).count() == 2

    def test_normaliza_el_telefono_al_formato_de_la_base(self, db, team):
        """Espacios, guiones, paréntesis y `00` se van; queda E.164."""
        contenido = hacer_excel([
            ENCABEZADOS[:2],
            ["Con guiones", "+57 300-111-2233"],
            ["Con paréntesis", "+52 (55) 1234 5678"],
            ["Con 00 internacional", "0057 3004445566"],
            ["Sin más, con país adentro", "573007778899"],
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert res.created == 4, res.errors
        guardados = {c.phone_e164 for c in db.query(models.Contact).all()}
        assert guardados == {
            "+573001112233",
            "+525512345678",
            "+573004445566",
            "+573007778899",
        }
        for numero in guardados:
            assert contactos_excel.E164_RE.match(numero), numero

    def test_las_columnas_extra_quedan_como_atributos(self, db, team):
        contenido = hacer_excel([
            ["Nombre", "Teléfono", "Ciudad", "Destino favorito"],
            ["Ana Ruiz", "+573001112233", "Cali", "Cartagena"],
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert res.created == 1
        assert sorted(res.detected_attributes) == ["Ciudad", "Destino favorito"]
        contacto = db.query(models.Contact).one()
        assert contacto.attributes == {"Ciudad": "Cali", "Destino favorito": "Cartagena"}
        # El encabezado se conserva tal como lo escribió la usuaria: es el
        # nombre que verá en el selector de variables de la campaña.
        assert "Destino favorito" in contacto.attributes

    def test_el_pais_por_defecto_completa_los_numeros_locales(self, db, team):
        contenido = hacer_excel([
            ENCABEZADOS[:2],
            ["Sin código", "300 111 2233"],
        ])

        res = contactos_excel.importar(db, team.id, contenido, prefijo_pais="57")

        assert res.created == 1, res.errors
        assert db.query(models.Contact).one().phone_e164 == "+573001112233"

    def test_respeta_el_opt_in_de_la_columna(self, db, team):
        contenido = hacer_excel([
            ENCABEZADOS[:4],
            ["Acepta", "+573001112233", "", "sí"],
            ["No acepta", "+573004445566", "", "no"],
            ["Sin decir nada", "+573007778899", "", ""],
        ])

        contactos_excel.importar(db, team.id, contenido)

        por_telefono = {c.phone_e164: c for c in db.query(models.Contact).all()}
        assert por_telefono["+573001112233"].opt_in is True
        assert por_telefono["+573004445566"].opt_in is False
        # Sin dato explícito, un contacto NUEVO nace con opt-in (mismo criterio
        # que el importador CSV que ya existía).
        assert por_telefono["+573007778899"].opt_in is True

    def test_una_columna_vacia_no_se_cuenta_como_fila(self, db, team):
        contenido = hacer_excel([
            ENCABEZADOS[:2],
            ["Ana", "+573001112233"],
            [None, None],
            ["", ""],
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert res.total == 1
        assert res.rejected == 0

    def test_ignora_la_fila_de_ejemplo_de_la_plantilla(self, db, team):
        """Si no borra el ejemplo, no le entra un contacto fantasma."""
        contenido = hacer_excel([
            contactos_excel.ENCABEZADOS_PLANTILLA,
            contactos_excel.FILA_EJEMPLO,
            ["Ana Ruiz", "+573001112233", "", "sí", "Cali", "Español"],
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert res.created == 1
        assert res.rejected == 1
        assert "ejemplo" in res.errors[0].reason
        assert db.query(models.Contact).count() == 1


# ---------------------------------------------------------------------------
# 2. Excel con teléfonos malos
# ---------------------------------------------------------------------------

class TestTelefonosMalos:
    def test_las_filas_buenas_entran_y_las_malas_se_reportan_con_su_fila(
        self, db, team
    ):
        contenido = hacer_excel([
            ENCABEZADOS[:2],
            ["Bien", "+573001112233"],        # fila 2 → ok
            ["Vacío", ""],                    # fila 3 → falta el teléfono
            ["Letras", "no tiene"],           # fila 4 → caracteres inválidos
            ["Cortísimo", "+12345"],          # fila 5 → muy pocos dígitos
            ["Sin país", "3001112233"],       # fila 6 → sin código de país
            ["Bien también", "+573004445566"],  # fila 7 → ok
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert (res.total, res.created, res.rejected) == (6, 2, 4)
        assert db.query(models.Contact).count() == 2
        filas_malas = [e.row for e in res.errors]
        assert filas_malas == [3, 4, 5, 6]
        motivos = {e.row: e.reason for e in res.errors}
        assert motivos[3] == "falta el teléfono"
        assert "código de país" in motivos[6]

    def test_el_motivo_nunca_lleva_el_telefono(self, db, team):
        """Regla 1: el mensaje ubica la fila, no repite el dato de la persona."""
        contenido = hacer_excel([
            ENCABEZADOS[:2],
            ["Malo", "3001112233"],
            ["Peor", "+00123"],
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert res.rejected == 2
        for error in res.errors:
            assert "3001112233" not in error.reason
            assert "00123" not in error.reason

    def test_correo_invalido_rechaza_la_fila_con_motivo_propio(self, db, team):
        contenido = hacer_excel([
            ENCABEZADOS[:3],
            ["Ana", "+573001112233", "esto no es un correo"],
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert res.rejected == 1
        assert "correo" in res.errors[0].reason
        assert db.query(models.Contact).count() == 0

    def test_archivo_sin_columna_de_telefono_no_revienta(self, db, team):
        contenido = hacer_excel([
            ["Nombre", "Ciudad"],
            ["Ana", "Cali"],
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert res.total == 0
        assert res.notice and "columna de teléfono" in res.notice
        assert db.query(models.Contact).count() == 0

    def test_archivo_que_no_es_excel_devuelve_aviso_y_no_excepcion(self, db, team):
        res = contactos_excel.importar(db, team.id, b"esto es un txt, no un xlsx")

        assert res.total == 0
        assert res.notice and ".xlsx" in res.notice

    @pytest.mark.parametrize(
        "bruto,esperado",
        [
            ("+573001112233", "+573001112233"),
            ("  +57 300 111 2233 ", "+573001112233"),
            ("57-300-111-2233", "+573001112233"),
            ("0057 300 111 2233", "+573001112233"),
            (573001112233, "+573001112233"),      # Excel lo pasa como número
            (573001112233.0, "+573001112233"),    # …o como float
        ],
    )
    def test_normalizar_telefono_acepta_lo_que_escribe_la_gente(self, bruto, esperado):
        assert contactos_excel.normalizar_telefono(bruto) == (esperado, None)

    @pytest.mark.parametrize(
        "bruto",
        ["", None, "abc", "+", "+0573001112233", "12345", "+5730011122334455667788"],
    )
    def test_normalizar_telefono_rechaza_lo_que_no_cumple_el_check(self, bruto):
        numero, motivo = contactos_excel.normalizar_telefono(bruto)
        assert numero is None
        assert motivo


# ---------------------------------------------------------------------------
# 3. Idempotencia
# ---------------------------------------------------------------------------

class TestIdempotencia:
    def test_reimportar_el_mismo_archivo_no_duplica(self, db, team):
        contenido = hacer_excel([
            ENCABEZADOS,
            ["Ana Ruiz", "+57 300 111 2233", "ana@ejemplo.com", "sí", "Cali", "Español"],
            ["Luis Paz", "+52 55 1234 5678", "", "sí", "CDMX", "Español"],
        ])

        primera = contactos_excel.importar(db, team.id, contenido)
        segunda = contactos_excel.importar(db, team.id, contenido)

        assert (primera.created, primera.updated) == (2, 0)
        assert (segunda.created, segunda.updated) == (0, 2)
        assert db.query(models.Contact).count() == 2

    def test_el_mismo_telefono_escrito_distinto_es_el_mismo_contacto(self, db, team):
        """`+57 300 111 2233` y `573001112233` son la misma persona."""
        primero = hacer_excel([ENCABEZADOS[:2], ["Ana", "+57 300 111 2233"]])
        segundo = hacer_excel([ENCABEZADOS[:2], ["Ana Ruiz", "573001112233"]])

        contactos_excel.importar(db, team.id, primero)
        res = contactos_excel.importar(db, team.id, segundo)

        assert (res.created, res.updated) == (0, 1)
        contacto = db.query(models.Contact).one()
        assert contacto.name == "Ana Ruiz"

    def test_reimportar_no_borra_atributos_que_no_vienen_en_el_archivo(
        self, db, team
    ):
        """Subir un Excel con menos columnas no debe vaciar lo ya guardado."""
        completo = hacer_excel([
            ["Nombre", "Teléfono", "Ciudad", "Idioma"],
            ["Ana", "+573001112233", "Cali", "Español"],
        ])
        parcial = hacer_excel([
            ["Nombre", "Teléfono", "Ciudad"],
            ["Ana", "+573001112233", "Bogotá"],
        ])

        contactos_excel.importar(db, team.id, completo)
        contactos_excel.importar(db, team.id, parcial)

        contacto = db.query(models.Contact).one()
        assert contacto.attributes == {"Ciudad": "Bogotá", "Idioma": "Español"}

    def test_reimportar_sin_columna_de_opt_in_no_reactiva_a_quien_dijo_que_no(
        self, db, team
    ):
        con_optin = hacer_excel([
            ENCABEZADOS[:4],
            ["Ana", "+573001112233", "", "no"],
        ])
        sin_optin = hacer_excel([
            ENCABEZADOS[:2],
            ["Ana", "+573001112233"],
        ])

        contactos_excel.importar(db, team.id, con_optin)
        contactos_excel.importar(db, team.id, sin_optin)

        assert db.query(models.Contact).one().opt_in is False

    def test_el_mismo_telefono_dos_veces_en_un_archivo_no_duplica(self, db, team):
        contenido = hacer_excel([
            ENCABEZADOS[:2],
            ["Ana", "+573001112233"],
            ["Ana otra vez", "+57 300 111 2233"],
        ])

        res = contactos_excel.importar(db, team.id, contenido)

        assert (res.created, res.updated) == (1, 1)
        assert db.query(models.Contact).count() == 1


# ---------------------------------------------------------------------------
# Aislamiento entre teams
# ---------------------------------------------------------------------------

def test_el_import_no_pisa_contactos_de_otro_team(db, team):
    """El mismo teléfono en dos teams son dos contactos distintos (S13-001)."""
    otro_usuario = models.User(
        nombre="Otra Agencia", tipo_documento="NIT", documento="900999",
        correo="otra@demo.com", hashed_password="x",
    )
    db.add(otro_usuario)
    db.commit()
    otro_team = models.Team(nombre="Otra Agencia", owner_user_id=otro_usuario.id)
    db.add(otro_team)
    db.commit()

    contenido = hacer_excel([ENCABEZADOS[:2], ["Ana", "+573001112233"]])
    contactos_excel.importar(db, team.id, contenido)
    res = contactos_excel.importar(db, otro_team.id, contenido)

    assert res.created == 1
    assert db.query(models.Contact).count() == 2


# ---------------------------------------------------------------------------
# Plantilla de guía
# ---------------------------------------------------------------------------

class TestPlantilla:
    def test_la_plantilla_tiene_encabezados_ejemplo_e_instrucciones(self):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(contactos_excel.construir_plantilla()))

        assert contactos_excel.HOJA_DATOS in wb.sheetnames
        assert contactos_excel.HOJA_INSTRUCCIONES in wb.sheetnames
        hoja = wb[contactos_excel.HOJA_DATOS]
        encabezados = [c.value for c in hoja[1]]
        assert encabezados == contactos_excel.ENCABEZADOS_PLANTILLA
        assert [c.value for c in hoja[2]] == contactos_excel.FILA_EJEMPLO
        guia = wb[contactos_excel.HOJA_INSTRUCCIONES]
        assert guia.cell(row=1, column=1).value

    def test_la_plantilla_recien_descargada_se_puede_volver_a_importar(
        self, db, team
    ):
        """El ciclo completo: descargo, lleno una fila, subo.

        Si la plantilla y el importador se desincronizan (un encabezado que el
        parser no reconoce), esta prueba es la que lo caza.
        """
        from openpyxl import Workbook, load_workbook

        wb = load_workbook(io.BytesIO(contactos_excel.construir_plantilla()))
        hoja = wb[contactos_excel.HOJA_DATOS]
        hoja.append(["Ana Ruiz", "+57 300 111 2233", "ana@ejemplo.com", "sí", "Cali", "Español"])
        buffer = io.BytesIO()
        wb.save(buffer)

        res = contactos_excel.importar(db, team.id, buffer.getvalue())

        assert res.created == 1, res.errors
        contacto = db.query(models.Contact).one()
        assert contacto.phone_e164 == "+573001112233"
        assert contacto.attributes == {"Ciudad": "Cali", "Idioma": "Español"}
        assert isinstance(Workbook(), Workbook)  # sanity del import


# ---------------------------------------------------------------------------
# Catálogo de campos
# ---------------------------------------------------------------------------

class TestCamposDisponibles:
    def test_siempre_ofrece_nombre_y_telefono(self, db, team):
        salida = contactos_excel.campos_disponibles(db, team.id)

        claves = [c.key for c in salida.fields]
        assert claves == ["name", "phone_e164"]
        assert salida.fields[0].token == "{{contact.name}}"

    def test_lista_los_atributos_de_los_contactos_del_team(self, db, team):
        contenido = hacer_excel([
            ["Nombre", "Teléfono", "Ciudad", "Destino favorito"],
            ["Ana", "+573001112233", "Cali", "Cartagena"],
            ["Luis", "+573004445566", "Bogotá", ""],
        ])
        contactos_excel.importar(db, team.id, contenido)

        salida = contactos_excel.campos_disponibles(db, team.id)

        atributos = {c.key: c for c in salida.fields if c.source == "attribute"}
        assert set(atributos) == {"Ciudad", "Destino favorito"}
        # Ordenados por cuántos contactos lo tienen (el más común primero).
        assert atributos["Ciudad"].contacts == 2
        assert atributos["Destino favorito"].contacts == 1
        assert atributos["Ciudad"].token == "{{contact.attributes.Ciudad}}"

    def test_no_ve_los_atributos_de_otro_team(self, db, team):
        ajeno = models.User(
            nombre="Ajena", tipo_documento="NIT", documento="900888",
            correo="ajena@demo.com", hashed_password="x",
        )
        db.add(ajeno)
        db.commit()
        otro_team = models.Team(nombre="Ajena", owner_user_id=ajeno.id)
        db.add(otro_team)
        db.commit()
        contactos_excel.importar(
            db,
            otro_team.id,
            hacer_excel([["Nombre", "Teléfono", "Secreto"], ["X", "+573001112233", "sí"]]),
        )

        salida = contactos_excel.campos_disponibles(db, team.id)

        assert all(c.key != "Secreto" for c in salida.fields)


# ---------------------------------------------------------------------------
# Los endpoints
# ---------------------------------------------------------------------------

@pytest.fixture
def cliente(Sesion, db, team):
    """TestClient sobre el router de contactos con la sesión del test.

    Se monta el router en una app propia (no `app.main`) porque ese módulo hace
    `create_all()` al importarse y exigiría un Postgres vivo. La membresía se
    inyecta ya resuelta: lo que se prueba aquí son las rutas, no el JWT.
    """
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    from app.dependencies import get_current_membership, get_db
    from app.routers import contacts as router_contacts

    miembro = models.TeamMember(
        team_id=team.id, user_id=team.owner_user_id, role="owner",
    )
    db.add(miembro)
    db.commit()

    app = FastAPI()
    app.include_router(router_contacts.router)

    def _get_db():
        sesion = Sesion()
        try:
            yield sesion
        finally:
            sesion.close()

    app.dependency_overrides[get_db] = _get_db
    app.dependency_overrides[get_current_membership] = lambda: miembro
    with TestClient(app) as c:
        yield c


class TestEndpoints:
    def test_la_plantilla_se_descarga_como_xlsx(self, cliente):
        """`/contacts/plantilla` no se confunde con `/contacts/{id}`.

        Si alguien mueve esta ruta debajo de la de detalle, FastAPI intenta
        parsear "plantilla" como un id entero y responde 422. Por eso se prueba.
        """
        res = cliente.get("/contacts/plantilla")

        assert res.status_code == 200, res.text
        assert res.headers["content-type"].startswith(contactos_excel.MIME_XLSX)
        assert "attachment" in res.headers["content-disposition"]
        assert res.content[:2] == b"PK"  # un .xlsx es un zip

    def test_import_excel_devuelve_el_resumen(self, cliente):
        contenido = hacer_excel([
            ENCABEZADOS,
            ["Ana Ruiz", "+57 300 111 2233", "", "sí", "Cali", "Español"],
            ["Malo", "no es un teléfono", "", "sí", "", ""],
        ])

        res = cliente.post(
            "/contacts/import-excel",
            files={"file": ("contactos.xlsx", contenido, contactos_excel.MIME_XLSX)},
        )

        assert res.status_code == 200, res.text
        cuerpo = res.json()
        assert cuerpo["created"] == 1
        assert cuerpo["rejected"] == 1
        assert cuerpo["errors"][0]["row"] == 3
        assert cuerpo["detected_attributes"] == ["Ciudad", "Idioma"]

    def test_import_excel_usa_el_pais_por_defecto_del_formulario(self, cliente):
        contenido = hacer_excel([ENCABEZADOS[:2], ["Ana", "3001112233"]])

        res = cliente.post(
            "/contacts/import-excel",
            files={"file": ("contactos.xlsx", contenido, contactos_excel.MIME_XLSX)},
            data={"pais_default": "+57"},
        )

        assert res.status_code == 200, res.text
        assert res.json()["created"] == 1

    def test_import_excel_rechaza_un_pais_por_defecto_absurdo(self, cliente):
        contenido = hacer_excel([ENCABEZADOS[:2], ["Ana", "3001112233"]])

        res = cliente.post(
            "/contacts/import-excel",
            files={"file": ("contactos.xlsx", contenido, contactos_excel.MIME_XLSX)},
            data={"pais_default": "colombia"},
        )

        assert res.status_code == 400
        assert "dígitos" in res.json()["detail"]

    def test_import_excel_rechaza_un_mime_que_no_es_excel(self, cliente):
        res = cliente.post(
            "/contacts/import-excel",
            files={"file": ("contactos.csv", b"a,b\n1,2", "text/csv")},
        )

        assert res.status_code == 415

    def test_import_excel_de_archivo_vacio(self, cliente):
        res = cliente.post(
            "/contacts/import-excel",
            files={"file": ("vacio.xlsx", b"", contactos_excel.MIME_XLSX)},
        )

        assert res.status_code == 400

    def test_campos_lista_nombre_telefono_y_atributos(self, cliente):
        contenido = hacer_excel([
            ["Nombre", "Teléfono", "Ciudad"],
            ["Ana", "+573001112233", "Cali"],
        ])
        cliente.post(
            "/contacts/import-excel",
            files={"file": ("c.xlsx", contenido, contactos_excel.MIME_XLSX)},
        )

        res = cliente.get("/contacts/campos")

        assert res.status_code == 200, res.text
        campos = res.json()["fields"]
        assert [c["key"] for c in campos] == ["name", "phone_e164", "Ciudad"]
        # Regla 2: sale el catálogo, no los valores. "Cali" no puede aparecer.
        assert "Cali" not in res.text

    def test_el_detalle_de_un_contacto_sigue_funcionando(self, cliente):
        """La ruta `/contacts/{id}` no se rompió al meter las nuevas."""
        contenido = hacer_excel([ENCABEZADOS[:2], ["Ana", "+573001112233"]])
        cliente.post(
            "/contacts/import-excel",
            files={"file": ("c.xlsx", contenido, contactos_excel.MIME_XLSX)},
        )
        listado = cliente.get("/contacts").json()

        res = cliente.get(f"/contacts/{listado[0]['id']}")

        assert res.status_code == 200
        assert res.json()["phone_e164"] == "+573001112233"
