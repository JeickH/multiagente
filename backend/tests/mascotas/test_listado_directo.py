"""Descarga del listado sin pasar por el bot (`GET /mascotas/listado/enlace`).

Por qué existe este atajo: de 75 conversaciones reales, 21 fueron solo para
bajar el Excel y 19 de ellas se resolvieron en un solo turno. Entregar un
archivo no necesita que nadie converse, y cada una de esas conversaciones
costaba una llamada completa al modelo (~US$0,03).

Lo que esta suite cuida es que el ahorro no se lleve por delante las dos reglas
del manual que protegen el archivo:

- **El enlace sigue firmado** (§2 del manual: nadie descarga por adivinar una
  URL) y es el mismo sobre Fernet que emite la herramienta del bot.
- **El listado sigue siendo solo de ENCONTRADAS**: los reportes de familias
  buscando llevan datos de contacto y no se reparten en un archivo.

Detalle de la suite: para descifrar el token se usa `router.decrypt_secret` y no
un `from app.services.crypto import ...` dentro del test. `tests/test_crypto.py`
recarga el módulo de cifrado con otra clave, así que un import tardío puede
quedarse con un Fernet distinto al que el router usó para firmar — y el test
falla por la clave, no por el código que dice estar probando.
"""
from __future__ import annotations

import json
from io import BytesIO

import pytest


ENLACE = "/mascotas/listado/enlace"


@pytest.fixture(autouse=True)
def limitador_limpio():
    """El rate-limit vive en el proceso, no en la base.

    `canal_limpio` (conftest) limpia los del chat y las fotos; este es el del
    listado. Sin esto, el test que llega al tope se lo deja puesto al siguiente.
    """
    from app.routers import mascotas as router

    def limpiar() -> None:
        router._listado_limiter._por_clave.clear()
        router._listado_limiter._todos.clear()

    limpiar()
    yield
    limpiar()


def _token_de(url: str) -> str:
    assert "/mascotas/listado.xlsx?token=" in url, url
    return url.split("token=", 1)[1]


def _sobre(url: str) -> dict:
    """Lo que lleva firmado el enlace."""
    from app.routers import mascotas as router

    return json.loads(router.decrypt_secret(_token_de(url)))


class TestEnlace:
    def test_lo_entrega_sin_gastar_una_llamada_al_modelo(self, cliente, respuestas):
        """El punto entero del atajo: descargar el listado no invoca a Bedrock."""
        modelo = respuestas()   # sin turnos guionizados: si lo llaman, revienta

        res = cliente.get(ENLACE)

        assert res.status_code == 200
        assert res.json()["filename"] == "mascotas_encontradas.xlsx"
        assert modelo.call_count == 0

    def test_no_pide_nombre_ni_telefono(self, cliente):
        """Es una descarga anónima: nadie tiene que devolverle la llamada a
        quien solo quiere el archivo, así que no se le piden datos."""
        assert cliente.get(ENLACE).status_code == 200

    def test_va_firmado(self, cliente):
        """El token es un sobre Fernet nuestro, no una URL adivinable."""
        sobre = _sobre(cliente.get(ENLACE).json()["url"])

        assert sobre["exp"] == "listado"

    def test_solo_autoriza_las_encontradas(self, cliente):
        """Manual §2: los reportes de quien busca llevan teléfono de familias y
        no se reparten en un archivo."""
        from app import models

        sobre = _sobre(cliente.get(ENLACE).json()["url"])

        assert sobre["t"] == models.MASCOTA_TIPO_ENCONTRADA

    def test_sin_base_publica_el_enlace_sale_por_el_rewrite_de_next(self, cliente):
        """En local el frontend llega al backend por `/api/*`; el enlace tiene
        que ser usable tal cual en un `href`, sin que la página lo arme."""
        url = cliente.get(ENLACE).json()["url"]

        assert url.startswith("/api/mascotas/listado.xlsx?token=")

    def test_en_produccion_apunta_a_la_base_publica(self, cliente, monkeypatch):
        """Igual que el enlace que manda el bot: absoluto contra la API."""
        monkeypatch.setenv("MASCOTAS_PUBLIC_BASE", "https://api.ejemplo.test/")

        url = cliente.get(ENLACE).json()["url"]

        assert url.startswith("https://api.ejemplo.test/mascotas/listado.xlsx?token=")

    def test_corta_el_abuso_con_rate_limit(self, cliente):
        """Es una descarga pública: se pide un par de veces, no doscientas."""
        from app.routers import mascotas as router

        for _ in range(router._listado_limiter._por_ip):
            assert cliente.get(ENLACE).status_code == 200

        res = cliente.get(ENLACE)
        assert res.status_code == 429
        # Mensaje genérico para el cliente (regla de seguridad #6).
        assert "listado" in res.json()["detail"]


class TestDescarga:
    def test_el_enlace_baja_el_excel(self, cliente, crear):
        """Round trip completo: el enlace que emite el endpoint sirve."""
        from openpyxl import load_workbook

        encontrada = crear(nombre="Rocky", raza="labrador")

        url = cliente.get(ENLACE).json()["url"]
        res = cliente.get(f"/mascotas/listado.xlsx?token={_token_de(url)}")

        assert res.status_code == 200
        assert "spreadsheetml" in res.headers["content-type"]
        assert "attachment" in res.headers["content-disposition"]
        hoja = load_workbook(BytesIO(res.content)).active
        celdas = {str(c.value) for fila in hoja.iter_rows() for c in fila}
        assert encontrada.codigo in celdas

    def test_no_reparte_los_reportes_de_quien_busca(self, cliente, crear):
        """Manual §2, la regla que de verdad importa aquí."""
        from openpyxl import load_workbook

        encontrada = crear(nombre="Rocky")
        perdida = crear(tipo_registro="perdida", nombre="Luna")

        url = cliente.get(ENLACE).json()["url"]
        res = cliente.get(f"/mascotas/listado.xlsx?token={_token_de(url)}")

        hoja = load_workbook(BytesIO(res.content)).active
        celdas = {str(c.value) for fila in hoja.iter_rows() for c in fila}
        assert encontrada.codigo in celdas
        assert perdida.codigo not in celdas

    def test_un_token_inventado_no_descarga_nada(self, cliente):
        assert cliente.get("/mascotas/listado.xlsx?token=no-soy-un-token").status_code == 403

    def test_un_sobre_nuestro_de_otra_cosa_tampoco(self, cliente):
        """Cifrar con nuestra clave no basta: el sobre tiene que decir que es
        para el listado. Evita que un token de sesión del chat sirva de pase."""
        from app.routers import mascotas as router

        ajeno = router.encrypt_secret(json.dumps({"h": [], "n": 0}))

        assert cliente.get(f"/mascotas/listado.xlsx?token={ajeno}").status_code == 403

    def test_la_descarga_tambien_tiene_rate_limit(self, cliente):
        """El token dura 24 h: sin tope, un enlace viejo alcanza para pedir el
        Excel indefinidamente, y armarlo lee la tabla entera."""
        from app.routers import mascotas as router

        token = _token_de(cliente.get(ENLACE).json()["url"])
        for _ in range(router._listado_limiter._por_ip - 1):
            assert cliente.get(f"/mascotas/listado.xlsx?token={token}").status_code == 200

        res = cliente.get(f"/mascotas/listado.xlsx?token={token}")
        assert res.status_code == 429
