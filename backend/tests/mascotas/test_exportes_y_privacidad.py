"""Exportes y privacidad: qué sale del sistema y con qué datos.

La regla que gobierna todo este archivo (manual §2, reglas 4 y 5): los
teléfonos son de familias que los dieron **para que las llamen por su
mascota**, no para redistribuirlos. Por eso:

- el Excel que reparte el bot es solo de **encontradas**;
- el JSON/ZIP para plataformas amigas sale **sin contacto** por defecto;
- la ficha que ve un desconocido en el chat **nunca** trae teléfono.
"""
from __future__ import annotations

import io
import json
import zipfile

import pytest

from app import models
from app.services import mascotas as svc

PERDIDA = {
    "tipo_registro": "perdida", "especie": "perro", "raza": "labrador",
    "nombre": "Lucas", "contacto_nombre": "Ana", "contacto_telefono": "3009998877",
    "ubicacion": "San Fernando, Cali",
}
ENCONTRADA = {
    "tipo_registro": "encontrada", "especie": "gato", "raza": "criollo",
    "contacto_nombre": "Pedro", "contacto_telefono": "3101112233",
    "ubicacion": "Guadalupe, Cali",
}


class TestFichaPublica:
    """Lo que el bot le puede mostrar a un desconocido."""

    def test_no_lleva_datos_de_contacto(self, db, crear):
        mascota = crear(**ENCONTRADA)
        ficha = svc.ficha_publica(mascota, db)

        assert "contacto_telefono" not in ficha
        assert "contacto_nombre" not in ficha
        assert "3101112233" not in json.dumps(ficha, ensure_ascii=False)
        assert "Pedro" not in json.dumps(ficha, ensure_ascii=False)

    def test_lleva_lo_que_sirve_para_reconocer_al_animal(self, db, crear):
        mascota = crear(**ENCONTRADA)
        ficha = svc.ficha_publica(mascota, db)
        for campo in ("codigo", "especie", "raza", "color", "zona", "fotos"):
            assert campo in ficha

    def test_dice_si_el_reporte_es_de_otra_plataforma(self, db, crear):
        """El bot lo necesita para saber si el contacto se entrega como
        teléfono (nuestro) o como enlace (importado)."""
        propio = crear(**ENCONTRADA)
        ajeno = crear(**{**ENCONTRADA, "origen_url": "https://encontradogs.co/pet/7"},
                      source="encontradogs")

        assert svc.ficha_publica(propio, db)["externo"] is False
        assert svc.ficha_publica(ajeno, db)["externo"] is True
        assert svc.ficha_publica(ajeno, db)["origen"] == "encontradogs"

    def test_la_zona_prefiere_el_barrio(self, db, crear):
        mascota = crear(**ENCONTRADA, barrio="Meléndez")
        assert svc.ficha_publica(mascota, db)["zona"] == "Meléndez"


class TestRepr:
    def test_el_repr_de_una_mascota_redacta_el_contacto(self, crear):
        """Regla de seguridad #1: nada de PII en los logs. Un `logger.info("%s",
        mascota)` no puede imprimir el teléfono de nadie."""
        mascota = crear(**PERDIDA)
        texto = repr(mascota)

        assert "3009998877" not in texto
        assert "REDACTED" in texto
        assert mascota.codigo in texto
        assert str(mascota) == repr(mascota)


class TestExcel:
    def _filas(self, contenido: bytes):
        from openpyxl import load_workbook

        ws = load_workbook(io.BytesIO(contenido)).active
        return list(ws.iter_rows(values_only=True))

    def test_el_listado_del_bot_es_solo_de_encontradas(self, db, crear):
        """Regla 5 del manual: los reportes de familias buscando llevan datos
        de contacto y no se reparten en un archivo."""
        crear(**PERDIDA)
        crear(**ENCONTRADA)

        filas = self._filas(svc.exportar_excel(db, tipo="encontrada"))

        codigos = [f[0] for f in filas[1:]]
        assert len(codigos) == 1
        assert "Mascota encontrada" in filas[1]
        # El de la familia que busca no aparece, ni su teléfono.
        assert "3009998877" not in str(filas)

    def test_el_export_del_panel_los_trae_todos(self, db, crear):
        crear(**PERDIDA)
        crear(**ENCONTRADA)
        filas = self._filas(svc.exportar_excel(db))
        assert len(filas) == 3      # encabezado + 2

    def test_tiene_encabezado_legible(self, db, crear):
        crear(**ENCONTRADA)
        filas = self._filas(svc.exportar_excel(db))
        assert filas[0][0] == "Código"
        assert "Teléfono" in filas[0]

    def test_los_reportes_cerrados_no_salen(self, db, crear):
        mascota = crear(**ENCONTRADA)
        mascota.estado = "reunida"
        db.commit()
        assert len(self._filas(svc.exportar_excel(db))) == 1   # solo encabezado

    def test_nombra_la_plataforma_de_origen(self, db, crear):
        crear(**{**ENCONTRADA, "origen_url": "https://patitasacasa.com/p/1"},
              source="patitasacasa")
        filas = self._filas(svc.exportar_excel(db))
        assert "Patitas a Casa" in str(filas[1])

    def test_lo_nuestro_se_nombra_como_la_iniciativa(self, db, crear):
        crear(**ENCONTRADA)
        assert "Recupera Tu Mascota" in str(self._filas(svc.exportar_excel(db))[1])


class TestJson:
    def test_por_defecto_va_sin_telefonos(self, db, crear):
        crear(**PERDIDA)
        crear(**ENCONTRADA)

        datos = svc.exportar_json(db)

        assert datos["contacto_incluido"] is False
        crudo = json.dumps(datos, ensure_ascii=False)
        assert "3009998877" not in crudo
        assert "3101112233" not in crudo
        assert "Ana" not in crudo

    def test_con_contacto_solo_si_se_pide_explicitamente(self, db, crear):
        crear(**PERDIDA)
        datos = svc.exportar_json(db, incluir_contacto=True)

        assert datos["contacto_incluido"] is True
        assert datos["casos"][0]["contacto_telefono"] == "3009998877"

    def test_cada_caso_dice_dónde_escribir(self, db, crear):
        """Sin teléfono, quien reconozca al animal tiene que saber a dónde ir."""
        propio = crear(**ENCONTRADA)
        ajeno = crear(**{**ENCONTRADA, "origen_url": "https://encontradogs.co/pet/7"},
                      source="encontradogs")

        por_codigo = {c["codigo"]: c for c in svc.exportar_json(db)["casos"]}

        assert por_codigo[propio.codigo]["contacto_url"] == \
            "https://mascotasperdidascolombia.com"
        assert por_codigo[ajeno.codigo]["contacto_url"] == \
            "https://encontradogs.co/pet/7"

    def test_lleva_version_y_licencia(self, db, crear):
        crear(**ENCONTRADA)
        datos = svc.exportar_json(db)
        assert datos["version"] == svc.EXPORT_VERSION
        assert "No usar con fines comerciales" in datos["licencia"]
        assert datos["total"] == len(datos["casos"])


class TestZip:
    def test_empaqueta_casos_leeme_y_fotos(self, db, crear):
        mascota = crear(**ENCONTRADA)
        foto = svc.guardar_foto(db, b"\xff\xd8imagen", "image/jpeg", mascota=mascota)

        paquete = svc.exportar_zip(db)

        with zipfile.ZipFile(io.BytesIO(paquete)) as zf:
            nombres = zf.namelist()
            assert "casos.json" in nombres
            assert "LEEME.txt" in nombres
            assert f"fotos/{mascota.codigo}/{foto.id}.jpg" in nombres
            casos = json.loads(zf.read("casos.json"))
        assert casos["total"] == 1

    def test_sin_fotos_si_no_se_piden(self, db, crear):
        mascota = crear(**ENCONTRADA)
        svc.guardar_foto(db, b"\xff\xd8imagen", "image/jpeg", mascota=mascota)

        with zipfile.ZipFile(io.BytesIO(svc.exportar_zip(db, incluir_fotos=False))) as zf:
            assert not [n for n in zf.namelist() if n.startswith("fotos/")]

    def test_el_leeme_advierte_cuando_lleva_telefonos(self, db, crear):
        crear(**PERDIDA)
        with zipfile.ZipFile(
            io.BytesIO(svc.exportar_zip(db, incluir_contacto=True))
        ) as zf:
            leeme = zf.read("LEEME.txt").decode()
        assert "datos personales" in leeme

    def test_sin_contacto_el_zip_no_tiene_telefonos(self, db, crear):
        crear(**PERDIDA)
        paquete = svc.exportar_zip(db)
        with zipfile.ZipFile(io.BytesIO(paquete)) as zf:
            todo = zf.read("casos.json").decode()
        assert "3009998877" not in todo


class TestResumen:
    def test_cuenta_por_tipo(self, db, crear):
        crear(**PERDIDA)
        crear(**ENCONTRADA)
        crear(**ENCONTRADA)

        resumen = svc.resumen(db)

        assert resumen["perdidas"] == 1
        assert resumen["encontradas"] == 2
